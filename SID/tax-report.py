"""SID offline tax-report tool.

Reads SID/closed-positions-sid.json (the authoritative record of every closed
SID trade with realised P&L) and writes an interactive Excel spreadsheet — ONE
row per closed trade — that a follower can hand to their accountant.

Runs entirely on the follower's OWN machine. The generated .xlsx is written to
SID/tax-reports/ which is git-ignored, so the sensitive tax document is NEVER
committed to the public repo (no private repo, no cloud artifact needed).

Columns (one row per trade):
    Symbol | Side | Opened | Closed | Quantity | Proceeds (USD) |
    Cost (USD) | Realised P&L (USD) | Month | UK Tax Year

  - Realised P&L uses the value already stored in the record (total_pnl /
    realizedPnl) as authoritative. Proceeds / Cost are shown for reference:
      long : cost     = entry * qty ,  proceeds = exit * qty
      short: proceeds = entry * qty ,  cost     = exit * qty
  - Month     = YYYY-MM of the CLOSE date.
  - UK Tax Year = tax year of the CLOSE date. The UK tax year runs 6 April to
    5 April, so e.g. a close on 2026-05-10 -> "2026/27"; 2026-03-20 ->
    "2025/26"; 2026-04-05 -> "2025/26"; 2026-04-06 -> "2026/27".

Interactivity:
  - AutoFilter on the header row (filter by Month / Tax Year / Symbol / Side).
  - A bold filter-aware TOTAL row using SUBTOTAL(109, ...) on the money columns.
    Function code 109 = SUM ignoring BOTH filtered-out AND manually-hidden rows,
    so "the total shows the visible rows only".

Output: SID/tax-reports/sid-tax-report-YYYY-MM-DD.xlsx (dir auto-created; date
from the system clock is fine — this runs on the user's machine, not in a
determinism-sensitive context).

Windows-safe: no emoji / unicode in print() (this repo has a cp1252 print trap).

Usage (or just double-click GENERATE-TAX-REPORT.bat / generate-tax-report.command):
    python SID/tax-report.py
"""
from __future__ import annotations

import io
import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] The 'openpyxl' library is not installed.")
    print("        Install it with:  python -m pip install openpyxl")
    print("        (Or run GENERATE-TAX-REPORT.bat / generate-tax-report.command,")
    print("         which installs it for you.)")
    sys.exit(1)

SID = Path(__file__).resolve().parent
CLOSED_PATH = SID / "closed-positions-sid.json"
OUT_DIR = SID / "tax-reports"

# ── Styling ──────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
NOTICE_FONT = Font(name="Calibri", size=10, italic=True, color="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E78")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")   # deep blue
HEADER_FILL = PatternFill("solid", fgColor="2F5F8F")  # mid blue
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")   # pale gold
THIN = Side(border_style="thin", color="B4B4B4")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTRE = Alignment(horizontal="center", vertical="center")

MONEY_FMT = '#,##0.00;[Red]-#,##0.00'
PNL_FMT = '+#,##0.00;[Red]-#,##0.00;0.00'
QTY_FMT = '#,##0'
DATE_FMT = 'yyyy-mm-dd'

NOTICE = (
    "SID realised-trades record for your tax return - NOT tax advice. "
    "Amounts in USD; your accountant converts to GBP at official HMRC rates. "
    "Filter the Month or 'UK Tax Year' column; the total row shows realised "
    "P&L for the visible rows."
)

# Column order (1-indexed positions used throughout).
COLUMNS = [
    "Symbol",
    "Side",
    "Opened",
    "Closed",
    "Quantity",
    "Proceeds (USD)",
    "Cost (USD)",
    "Realised P&L (USD)",
    "Month",
    "UK Tax Year",
]
COL = {name: i + 1 for i, name in enumerate(COLUMNS)}


# ── Helpers (defensive field access) ─────────────────────────────────────────
def _first(rec: dict, *keys):
    """Return the first present, non-None value among keys, else None."""
    for k in keys:
        if k in rec and rec[k] is not None:
            return rec[k]
    return None


def _to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _parse_date(val):
    """Parse a date-ish value into a datetime.date, or None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s:
        return None
    # Take the date part if a full ISO timestamp slipped in.
    s = s.replace("T", " ").split(" ")[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def uk_tax_year(d: date) -> str:
    """UK tax year label for a date. Tax year runs 6 April to 5 April.

    On or after 6 April -> that calendar year starts the tax year.
    On or before 5 April -> the previous calendar year started it.
    e.g. 2026-05-10 -> '2026/27'; 2026-03-20 -> '2025/26';
         2026-04-05 -> '2025/26'; 2026-04-06 -> '2026/27'.
    """
    if (d.month, d.day) >= (4, 6):
        start = d.year
    else:
        start = d.year - 1
    return f"{start}/{str((start + 1) % 100).zfill(2)}"


def build_row(rec: dict):
    """Turn a closed-position record into a display row, or (None, reason)."""
    symbol = _first(rec, "symbol")
    side = _first(rec, "side")
    if symbol is None:
        return None, "missing symbol"

    side_str = str(side).lower() if side is not None else ""

    opened = _parse_date(_first(rec, "openDate", "open_date", "signalDate"))
    closed = _parse_date(_first(rec, "closeDate", "close_date"))
    if closed is None:
        return None, f"{symbol}: missing/unparseable close date"

    # Quantity: prefer the total position size actually traded.
    qty = _to_float(
        _first(rec, "shares_total", "exit_shares", "shares", "shares_remaining")
    )
    if qty is None:
        qty = 0.0

    # Authoritative realised P&L already stored in the record.
    pnl = _to_float(_first(rec, "total_pnl", "realizedPnl", "realized_pnl", "exit_pnl"))
    if pnl is None:
        return None, f"{symbol}: no realised P&L field (total_pnl/realizedPnl)"

    # Prices for Proceeds / Cost (reference only).
    entry = _to_float(_first(rec, "entry", "entryPrice", "entry_price"))
    exit_px = _to_float(
        _first(rec, "exit_price", "exitPrice", "exit_avg_price", "tp1_price")
    )

    proceeds = None
    cost = None
    if entry is not None and exit_px is not None and qty:
        if side_str == "short":
            proceeds = entry * qty     # sold first (at entry)
            cost = exit_px * qty       # bought back (at exit) to cover
        else:  # long (default)
            cost = entry * qty         # bought first (at entry)
            proceeds = exit_px * qty   # sold (at exit)

    month = f"{closed.year:04d}-{closed.month:02d}"
    tax_yr = uk_tax_year(closed)

    return {
        "Symbol": str(symbol),
        "Side": side_str or "?",
        "Opened": opened,
        "Closed": closed,
        "Quantity": qty,
        "Proceeds (USD)": proceeds,
        "Cost (USD)": cost,
        "Realised P&L (USD)": pnl,
        "Month": month,
        "UK Tax Year": tax_yr,
    }, None


def load_closed():
    if not CLOSED_PATH.exists():
        print("[WARN] " + str(CLOSED_PATH) + " not found.")
        print("       This is normal if your bot has not closed any trades yet.")
        return []
    try:
        with io.open(CLOSED_PATH, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (ValueError, OSError) as exc:
        print("[ERROR] Could not read " + str(CLOSED_PATH) + ": " + str(exc))
        return []
    if not isinstance(data, list):
        print("[ERROR] Expected a JSON array in closed-positions-sid.json.")
        return []
    return data


def build_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SID Realised Trades"

    n_cols = len(COLUMNS)
    last_col_letter = get_column_letter(n_cols)

    # Row 1: notice banner (merged across all columns).
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    banner = ws.cell(row=1, column=1)
    banner.value = NOTICE
    banner.font = NOTICE_FONT
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 42

    # Row 2: title.
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    title = ws.cell(row=2, column=1)
    title.value = "SID - Realised Trades for Tax"
    title.font = TITLE_FONT
    title.fill = TITLE_FILL
    title.alignment = ALIGN_LEFT

    # Row 3: header.
    header_row = 3
    for name, idx in COL.items():
        c = ws.cell(row=header_row, column=idx)
        c.value = name
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTRE
        c.border = BORDER_THIN

    data_start = header_row + 1
    r = data_start
    for row in rows:
        ws.cell(row=r, column=COL["Symbol"], value=row["Symbol"]).alignment = ALIGN_LEFT
        ws.cell(row=r, column=COL["Side"], value=row["Side"]).alignment = ALIGN_CENTRE

        c_open = ws.cell(row=r, column=COL["Opened"], value=row["Opened"])
        c_open.number_format = DATE_FMT
        c_open.alignment = ALIGN_CENTRE
        c_close = ws.cell(row=r, column=COL["Closed"], value=row["Closed"])
        c_close.number_format = DATE_FMT
        c_close.alignment = ALIGN_CENTRE

        c_qty = ws.cell(row=r, column=COL["Quantity"], value=row["Quantity"])
        c_qty.number_format = QTY_FMT
        c_qty.alignment = ALIGN_CENTRE

        c_proc = ws.cell(row=r, column=COL["Proceeds (USD)"], value=row["Proceeds (USD)"])
        c_proc.number_format = MONEY_FMT
        c_cost = ws.cell(row=r, column=COL["Cost (USD)"], value=row["Cost (USD)"])
        c_cost.number_format = MONEY_FMT

        c_pnl = ws.cell(row=r, column=COL["Realised P&L (USD)"], value=row["Realised P&L (USD)"])
        c_pnl.number_format = PNL_FMT

        ws.cell(row=r, column=COL["Month"], value=row["Month"]).alignment = ALIGN_CENTRE
        ws.cell(row=r, column=COL["UK Tax Year"], value=row["UK Tax Year"]).alignment = ALIGN_CENTRE
        r += 1

    data_end = r - 1                 # last data row (== header row if zero trades)
    has_data = data_end >= data_start

    # ── Filter-aware TOTAL row (SUBTOTAL 109 = SUM ignoring hidden rows) ──────
    total_row = (data_end if has_data else header_row) + 1
    tc = ws.cell(row=total_row, column=COL["Symbol"])
    tc.value = "REALISED P&L (visible rows)"

    if has_data:
        proc_col = get_column_letter(COL["Proceeds (USD)"])
        cost_col = get_column_letter(COL["Cost (USD)"])
        pnl_col = get_column_letter(COL["Realised P&L (USD)"])
        ws.cell(row=total_row, column=COL["Proceeds (USD)"]).value = (
            f"=SUBTOTAL(109,{proc_col}{data_start}:{proc_col}{data_end})"
        )
        ws.cell(row=total_row, column=COL["Cost (USD)"]).value = (
            f"=SUBTOTAL(109,{cost_col}{data_start}:{cost_col}{data_end})"
        )
        ws.cell(row=total_row, column=COL["Realised P&L (USD)"]).value = (
            f"=SUBTOTAL(109,{pnl_col}{data_start}:{pnl_col}{data_end})"
        )
    else:
        ws.cell(row=total_row, column=COL["Proceeds (USD)"]).value = 0
        ws.cell(row=total_row, column=COL["Cost (USD)"]).value = 0
        ws.cell(row=total_row, column=COL["Realised P&L (USD)"]).value = 0

    ws.cell(row=total_row, column=COL["Proceeds (USD)"]).number_format = MONEY_FMT
    ws.cell(row=total_row, column=COL["Cost (USD)"]).number_format = MONEY_FMT
    ws.cell(row=total_row, column=COL["Realised P&L (USD)"]).number_format = PNL_FMT
    for col_idx in range(1, n_cols + 1):
        tcell = ws.cell(row=total_row, column=col_idx)
        tcell.font = TOTAL_FONT
        tcell.fill = TOTAL_FILL

    # ── AutoFilter on the header + data range (or just the header) ───────────
    filter_end = data_end if has_data else header_row
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{filter_end}"

    # Freeze so the header stays put while scrolling / filtering.
    ws.freeze_panes = f"A{data_start}"

    # Column widths.
    widths = {
        "Symbol": 10, "Side": 8, "Opened": 12, "Closed": 12, "Quantity": 10,
        "Proceeds (USD)": 15, "Cost (USD)": 15, "Realised P&L (USD)": 18,
        "Month": 10, "UK Tax Year": 12,
    }
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(COL[name])].width = w

    return wb


def main():
    print("[SID] Building your realised-trades tax report...")
    records = load_closed()

    rows = []
    skipped = 0
    for rec in records:
        row, reason = build_row(rec)
        if row is None:
            skipped += 1
            print("[SKIP] " + str(reason))
            continue
        rows.append(row)

    # Sort by close date so the report reads chronologically.
    rows.sort(key=lambda x: (x["Closed"], x["Symbol"]))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = date.today().strftime("%Y-%m-%d")
    out_path = OUT_DIR / f"sid-tax-report-{stamp}.xlsx"

    wb = build_workbook(rows)
    wb.save(out_path)

    print("[OK] Included " + str(len(rows)) + " closed trade(s).")
    if skipped:
        print("[OK] Skipped " + str(skipped) + " malformed/incomplete record(s) (see [SKIP] above).")
    print("[DONE] Report saved to:")
    print("       " + str(out_path))
    print("       (This file stays on YOUR computer - nothing is uploaded.)")
    # Print the path on its own last line so launchers can capture it if needed.
    return 0


if __name__ == "__main__":
    sys.exit(main())
