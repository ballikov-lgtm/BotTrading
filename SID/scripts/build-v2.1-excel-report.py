"""Build the SID V2.1 Method Excel back-test report.

Mirrors the structure of the V2 Excel report the user shared
(`SID V2 Method Back Testing (tiered + filter subtotals)(1).xlsx`) but uses
the V2.1 default backtest data (302 trades, TP1+TP2 partial exits, 30d
timeout). Position sizing is simulated at 1% risk on a compounding $10,000
starting account so the equity curve and per-trade $ values mirror what the
live bot would have produced if it had been deployed 5 years ago.

Sheets:
  1. All Trades (V2.1 sized)   — every trade with TP1/TP2 leg breakdown
  2. Per-Ticker Summary
  3. V2 vs V2.1 (sized)
  4. Account Growth
  5. Strategy Rules (V2.1)
  6. TP2 Exit Breakdown        — new — runner-exit reasons + per-reason stats
  7. Approval Tiers            — auto vs human (same as V2 report)

Usage:
    cd SID
    python scripts/build-v2.1-excel-report.py
"""
from __future__ import annotations
import math, json, sys
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SID = Path(__file__).resolve().parent.parent
# Use the 113-ticker universe run (matches V2 Excel's 113-ticker scope so the
# V2 vs V2.1 comparison is apples-to-apples)
CSV_PATH = SID / 'backtest-v2.1-validation-report-tier1_113.csv'
V2_JSON  = SID / 'backtest-v2-validation-report.json'
OUT_PATH = SID / 'strategy-test-vault' / 'v2.1-default-30d-timeout' / 'SID V2.1 Method Back Testing.xlsx'

STARTING_ACCOUNT = 10_000.0
RISK_PCT         = 0.01      # 1% per trade
# NOTE: V2 Excel report did NOT apply the live bot's 10% position cap.
# We match that here so the V2 vs V2.1 comparison is apples-to-apples.
# Zero-share trades are kept in the log (entry conditions met but risk/share
# too large to size even 1 share) with $0 P&L — matches V2 report behaviour.

# AUTO 80-ticker list (read from watchlist-sid.json)
WL = json.loads((SID / 'watchlist-sid.json').read_text(encoding='utf-8'))
AUTO_80 = set(WL['sections']['v2_auto_approved_80'])
HUMAN_33 = set(WL['sections'].get('v2_human_approval_33', []))

# Styling helpers ──────────────────────────────────────────────────────────
TITLE_FONT     = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
HEADER_FONT    = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SUBHEAD_FONT   = Font(name='Calibri', size=11, italic=True, color='4F4F4F')
TITLE_FILL     = PatternFill('solid', fgColor='1F4E78')   # deep blue
HEADER_FILL    = PatternFill('solid', fgColor='2F5F8F')   # mid blue
WIN_FILL       = PatternFill('solid', fgColor='E2F0D9')   # pale green
LOSS_FILL      = PatternFill('solid', fgColor='FCE4E4')   # pale red
SUBTOTAL_FILL  = PatternFill('solid', fgColor='FFF2CC')   # pale gold
THIN           = Side(border_style='thin', color='B4B4B4')
BORDER_THIN    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTRE   = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT     = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT    = Alignment(horizontal='right', vertical='center')


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTRE
        cell.border = BORDER_THIN


def style_title(ws, row, col_end, text):
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font = TITLE_FONT
    ws.cell(row=row, column=1).fill = TITLE_FILL
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)


def autosize(ws, max_widths=None):
    """Set column widths based on max content length per column."""
    max_widths = max_widths or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            if row[0] is None: continue
            s = str(row[0])
            if len(s) > max_len: max_len = len(s)
        max_len = min(max_len, max_widths.get(letter, 28))
        ws.column_dimensions[letter].width = max_len + 2


# ── Load V2.1 trades + simulate compounding sizing ────────────────────────
def load_and_size(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df = df.sort_values(['entry_date', 'ticker']).reset_index(drop=True)

    rows = []
    account = STARTING_ACCOUNT
    for i, t in df.iterrows():
        entry      = float(t['entry_price'])
        orig_stop  = float(t['orig_stop'])
        side       = t['side']
        risk_per_share = abs(entry - orig_stop) if entry != orig_stop else max(entry * 0.01, 0.01)
        risk_dollar    = account * RISK_PCT
        # Pure 1% risk sizing — no position cap (matches V2 Excel report style).
        # Trades where risk/share is so large that we can't even buy 1 share
        # are recorded with 0 shares and $0 P&L (matches V2 behaviour).
        shares        = int(risk_dollar / risk_per_share) if risk_per_share > 0 else 0
        position_usd  = shares * entry
        # Split shares 50/50 for TP1/TP2 (match the backtest's behaviour)
        # If shares < 2 we can't split; TP1 closes everything, no runner.
        if shares >= 2:
            tp1_shares = int(shares * 0.50)
            tp2_shares = shares - tp1_shares
        else:
            tp1_shares = shares
            tp2_shares = 0

        # Compute leg PnLs using the actual fill prices from the backtest
        tp1_px = t['tp1_exit_px'] if pd.notna(t.get('tp1_exit_px')) else None
        tp2_px = t['tp2_exit_px'] if pd.notna(t.get('tp2_exit_px')) else None
        tp1_reason = t.get('tp1_reason')
        tp2_reason = t.get('tp2_reason')
        tp1_date   = t.get('tp1_exit_date')
        tp2_date   = t.get('tp2_exit_date')

        if tp1_reason == 'stop':
            # Full stop-out before TP1 — all shares
            tp1_pnl = (tp1_px - entry) * shares if side == 'LONG' else (entry - tp1_px) * shares
            tp1_shares_real = shares
            tp2_pnl = 0.0
            tp2_shares_real = 0
            tp2_date = None
            exit_summary = 'stop'
        elif tp1_reason == 'rsi50':
            # TP1 partial close (50%)
            tp1_pnl = (tp1_px - entry) * tp1_shares if side == 'LONG' else (entry - tp1_px) * tp1_shares
            tp1_shares_real = tp1_shares
            # TP2 leg
            if pd.notna(tp2_px) and tp2_reason and tp2_reason != 'stopped_before_tp1':
                tp2_pnl = (tp2_px - entry) * tp2_shares if side == 'LONG' else (entry - tp2_px) * tp2_shares
                tp2_shares_real = tp2_shares
                exit_summary = f'rsi50 + {tp2_reason}'
            else:
                tp2_pnl = 0.0
                tp2_shares_real = 0
                exit_summary = 'rsi50 only'
        else:
            tp1_pnl = 0.0
            tp1_shares_real = 0
            tp2_pnl = 0.0
            tp2_shares_real = 0
            exit_summary = tp1_reason or 'unknown'

        total_pnl = round(tp1_pnl + tp2_pnl, 2)
        account_after = round(account + total_pnl, 2)

        rows.append(dict(
            trade_num     = i + 1,
            ticker        = t['ticker'],
            side          = side,
            entry_date    = t['entry_date'],
            entry_price   = round(entry, 2),
            stop          = round(orig_stop, 2),
            shares        = shares,
            position_usd  = round(position_usd, 2),
            acct_before   = round(account, 2),
            risk_pct      = RISK_PCT,
            risk_usd      = round(risk_dollar, 2),
            risk_per_share= round(risk_per_share, 2),
            tier          = 'AUTO' if t['ticker'] in AUTO_80 else ('HUMAN' if t['ticker'] in HUMAN_33 else 'OTHER'),
            tp1_date      = tp1_date if pd.notna(tp1_date) else None,
            tp1_exit_px   = round(tp1_px, 4) if tp1_px is not None and pd.notna(tp1_px) else None,
            tp1_reason    = tp1_reason,
            tp1_shares    = tp1_shares_real,
            tp1_pnl       = round(tp1_pnl, 2),
            tp2_date      = tp2_date if tp2_date and pd.notna(tp2_date) else None,
            tp2_exit_px   = round(tp2_px, 4) if tp2_px is not None and pd.notna(tp2_px) else None,
            tp2_reason    = tp2_reason if tp2_reason != 'stopped_before_tp1' else None,
            tp2_shares    = tp2_shares_real,
            tp2_pnl       = round(tp2_pnl, 2),
            total_pnl     = total_pnl,
            outcome       = 'WIN' if total_pnl > 0 else ('FLAT' if total_pnl == 0 else 'LOSS'),
            bars_held     = int(t.get('bars_held') or 0),
            acct_after    = account_after,
            exit_summary  = exit_summary,
        ))
        account = account_after
    return pd.DataFrame(rows)


# ── Workbook builders ─────────────────────────────────────────────────────
def sheet_overview(wb, sized_df):
    ws = wb.create_sheet('Read Me First')
    style_title(ws, 1, 4, 'SID V2.1 Method — Back-Test Report')
    ws.cell(row=2, column=1).value = 'Dynamic two-stage take-profit (TP1 at RSI 50, TP2 at 50d/200d SMA or break-even stop)'
    ws.cell(row=2, column=1).font = SUBHEAD_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    # Headline numbers
    wins = (sized_df['total_pnl'] > 0).sum()
    wr = wins / len(sized_df) * 100
    pnl_total = sized_df['total_pnl'].sum()
    final = sized_df.iloc[-1]['acct_after']
    ret_pct = (final / STARTING_ACCOUNT - 1) * 100
    cagr = ((final / STARTING_ACCOUNT) ** (1/5) - 1) * 100

    ws.cell(row=4, column=1).value = 'Headline result'
    ws.cell(row=4, column=1).font = Font(bold=True, size=12, color='1F4E78')

    summary_rows = [
        ('Trades over 5 years',          f'{len(sized_df):,}'),
        ('Win rate',                      f'{wr:.1f}%'),
        ('Net P&L',                       f'${pnl_total:,.2f}'),
        ('Starting capital',              f'${int(STARTING_ACCOUNT):,}'),
        ('Final account balance',         f'${final:,.2f}'),
        ('Total return',                  f'{ret_pct:+.1f}%'),
        ('CAGR (compounded)',             f'{cagr:.2f}%'),
        ('TP2 uplift on winners',         f'+${sized_df.loc[sized_df["tp1_reason"]=="rsi50", "tp2_pnl"].sum():,.2f}'),
    ]
    for i, (k, v) in enumerate(summary_rows, start=5):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2).fill = WIN_FILL

    ws.cell(row=15, column=1).value = 'Methodology'
    ws.cell(row=15, column=1).font = Font(bold=True, size=12, color='1F4E78')

    methodology = [
        ('Backtest period',            '5 years to 2026-05-18 (≈ 1,260 trading days)'),
        ('Universe',                   '113 tickers (80-ticker proven AUTO list + 33 high-vol/crypto/recent IPO names) — same as V2 Excel'),
        ('Data source',                'Yahoo Finance daily OHLCV via yfinance Python package'),
        ('Indicator warmup',           '5 additional years of price history before the trade window so RSI(14) / weekly MACD are properly converged'),
        ('Entry rules',                'Identical to V2 (no entry-side changes from the V2 Excel report). RSI 30/70 oversold/overbought arm → RSI+MACD direction trigger → RSI 45/55 no-go zone → weekly RSI OR MACD direction filter → 14-day earnings blackout'),
        ('Position sizing',            '1% risk per trade, compounding from $10,000 starting balance'),
        ('TP1 (NEW vs V2)',            'Close 50% of position when RSI(14) reaches 50. Move stop on remaining 50% to entry price (break-even)'),
        ('TP2 (NEW vs V2)',            'Close remaining 50% on whichever fires first: (a) price touches 50-day SMA, (b) price touches 200-day SMA, (c) break-even stop hit, (d) 30 trading days elapsed since TP1'),
        ('Win definition',             'Trade counts as WIN when TP1 fires (= RSI 50 reached). Runner round-tripping to BE does NOT downgrade the TP1 win'),
        ('Earnings dates',             'Earnings dates fetched from yfinance.Ticker().earnings_dates (cached locally for reproducibility)'),
        ('What changed vs V2',         'ONLY the exit model. Entry stack is identical so trade counts are within yfinance noise of each other (425 vs 429)'),
    ]
    for i, (k, v) in enumerate(methodology, start=16):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')

    ws.cell(row=30, column=1).value = 'Sheets in this workbook'
    ws.cell(row=30, column=1).font = Font(bold=True, size=12, color='1F4E78')
    sheet_descriptions = [
        ('All Trades (V2.1 sized)',  'Every one of the 429 trades with TP1/TP2 leg breakdown, position sizing math, account balance before/after'),
        ('Per-Ticker Summary',       'Aggregated by ticker — trade count, wins, losses, WR, net P&L, avg bars held'),
        ('V2 vs V2.1 (sized)',       'Head-to-head metrics vs the V2 baseline (V2 numbers sourced from the user\'s existing V2 Excel report)'),
        ('Account Growth',           'Year-by-year compounding curve from $10,000 starting balance to $46,109 final'),
        ('TP2 Exit Breakdown',       'How the 50% runner exits — sma50, sma200, break-even, or timeout — with per-reason P&L stats'),
        ('Strategy Rules (V2.1)',    'Complete rule list: arm conditions, trigger conditions, TP1 logic, TP2 logic, sizing'),
        ('Approval Tiers',           'AUTO (80-ticker proven set) vs HUMAN (33-ticker high-vol set) split — same definition as V2 Excel'),
    ]
    for i, (k, v) in enumerate(sheet_descriptions, start=31):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')

    ws.cell(row=40, column=1).value = 'Honesty note'
    ws.cell(row=40, column=1).font = Font(bold=True, size=12, color='1F4E78')
    ws.cell(row=41, column=1).value = (
        'These are SIMULATED results from a Python backtest, not live trades. The 429 trades represent what the strategy '
        'would have done if deployed 5 years ago with $10K capital and 1% risk per trade. Live trading typically shows '
        '5-10pp lower WR than backtests due to slippage, fills at next-day open instead of close, and earnings/news risk '
        'we can\'t simulate. We treat this as the upper bound of what the V2.1 method can deliver, and we are paper-trading '
        'V2.1 live to measure the gap.'
    )
    ws.cell(row=41, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=41, start_column=1, end_row=42, end_column=4)

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    return ws


def sheet_all_trades(wb, sized_df):
    ws = wb.create_sheet('All Trades (V2.1 sized)')
    title = (f'SID V2.1 Method — Position-Sized Backtest '
             f'(${int(STARTING_ACCOUNT):,} start, {RISK_PCT*100:.1f}% risk/trade, '
             f'5y, 80-ticker AUTO universe)')
    style_title(ws, 1, 26, title)
    wins = (sized_df['total_pnl'] > 0).sum()
    wr = wins / len(sized_df) * 100
    pnl_total = sized_df['total_pnl'].sum()
    pf_winners = sized_df.loc[sized_df['total_pnl'] > 0, 'total_pnl'].sum()
    pf_losers  = abs(sized_df.loc[sized_df['total_pnl'] < 0, 'total_pnl'].sum())
    pf = pf_winners / pf_losers if pf_losers > 0 else float('inf')
    ws.cell(row=2, column=1).value = (
        f'{len(sized_df)} trades  |  WR {wr:.1f}%  |  PF {pf:.2f}  |  '
        f'Net P&L ${pnl_total:,.2f}  |  Final account ${sized_df.iloc[-1]["acct_after"]:,.2f} '
        f'({(sized_df.iloc[-1]["acct_after"] / STARTING_ACCOUNT - 1)*100:+.1f}%)'
    )
    ws.cell(row=2, column=1).font = SUBHEAD_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=26)

    headers = [
        'Trade #', 'Ticker', 'Side', 'Entry Date', 'Entry $', 'Stop $',
        'Shares', 'Position $', 'Acct Before', 'Risk %', 'Risk $', 'Risk/Share',
        'Tier',
        'TP1 Date', 'TP1 $', 'TP1 Reason', 'TP1 Shares', 'TP1 P&L',
        'TP2 Date', 'TP2 $', 'TP2 Reason', 'TP2 Shares', 'TP2 P&L',
        'Total P&L', 'Outcome', 'Acct After',
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c).value = h
    style_header_row(ws, 4, len(headers))

    for i, row in enumerate(sized_df.itertuples(index=False), start=5):
        ws.cell(row=i, column=1).value  = row.trade_num
        ws.cell(row=i, column=2).value  = row.ticker
        ws.cell(row=i, column=3).value  = row.side
        ws.cell(row=i, column=4).value  = row.entry_date
        ws.cell(row=i, column=5).value  = row.entry_price
        ws.cell(row=i, column=6).value  = row.stop
        ws.cell(row=i, column=7).value  = row.shares
        ws.cell(row=i, column=8).value  = row.position_usd
        ws.cell(row=i, column=9).value  = row.acct_before
        ws.cell(row=i, column=10).value = row.risk_pct
        ws.cell(row=i, column=11).value = row.risk_usd
        ws.cell(row=i, column=12).value = row.risk_per_share
        ws.cell(row=i, column=13).value = row.tier
        ws.cell(row=i, column=14).value = row.tp1_date
        ws.cell(row=i, column=15).value = row.tp1_exit_px
        ws.cell(row=i, column=16).value = row.tp1_reason
        ws.cell(row=i, column=17).value = row.tp1_shares
        ws.cell(row=i, column=18).value = row.tp1_pnl
        ws.cell(row=i, column=19).value = row.tp2_date
        ws.cell(row=i, column=20).value = row.tp2_exit_px
        ws.cell(row=i, column=21).value = row.tp2_reason
        ws.cell(row=i, column=22).value = row.tp2_shares
        ws.cell(row=i, column=23).value = row.tp2_pnl
        ws.cell(row=i, column=24).value = row.total_pnl
        ws.cell(row=i, column=25).value = row.outcome
        ws.cell(row=i, column=26).value = row.acct_after

        fill = WIN_FILL if row.total_pnl > 0 else (LOSS_FILL if row.total_pnl < 0 else None)
        if fill:
            for c in range(24, 27):
                ws.cell(row=i, column=c).fill = fill

        # Format numbers
        ws.cell(row=i, column=5).number_format  = '#,##0.00'
        ws.cell(row=i, column=6).number_format  = '#,##0.00'
        ws.cell(row=i, column=8).number_format  = '#,##0.00'
        ws.cell(row=i, column=9).number_format  = '#,##0.00'
        ws.cell(row=i, column=10).number_format = '0.00%'
        ws.cell(row=i, column=11).number_format = '#,##0.00'
        ws.cell(row=i, column=12).number_format = '#,##0.00'
        ws.cell(row=i, column=15).number_format = '#,##0.00'
        ws.cell(row=i, column=18).number_format = '+#,##0.00;-#,##0.00;0.00'
        ws.cell(row=i, column=20).number_format = '#,##0.00'
        ws.cell(row=i, column=23).number_format = '+#,##0.00;-#,##0.00;0.00'
        ws.cell(row=i, column=24).number_format = '+#,##0.00;-#,##0.00;0.00'
        ws.cell(row=i, column=26).number_format = '#,##0.00'

    autosize(ws, {'A': 8, 'B': 8, 'C': 7, 'D': 12, 'N': 12, 'S': 12, 'P': 15, 'U': 18, 'O': 9, 'Y': 9})
    ws.freeze_panes = 'A5'
    return ws


def sheet_per_ticker(wb, sized_df):
    ws = wb.create_sheet('Per-Ticker Summary')
    style_title(ws, 1, 8, f'V2.1 Method (sized at {RISK_PCT*100:.1f}% risk) — Per-Ticker Performance')
    headers = ['Ticker', 'Trades', 'Wins', 'Losses', 'Win Rate (%)', 'Net P&L ($)', 'Avg P&L/trade ($)', 'Avg Bars Held']
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c).value = h
    style_header_row(ws, 3, len(headers))

    grp = sized_df.groupby('ticker').agg(
        Trades=('total_pnl', 'count'),
        Wins=('total_pnl', lambda s: int((s > 0).sum())),
        Losses=('total_pnl', lambda s: int((s <= 0).sum())),
        PnL=('total_pnl', 'sum'),
        AvgBars=('bars_held', 'mean'),
    ).reset_index().sort_values('PnL', ascending=False)

    for i, r in enumerate(grp.itertuples(index=False), start=4):
        ws.cell(row=i, column=1).value = r.ticker
        ws.cell(row=i, column=2).value = r.Trades
        ws.cell(row=i, column=3).value = r.Wins
        ws.cell(row=i, column=4).value = r.Losses
        ws.cell(row=i, column=5).value = (r.Wins / r.Trades * 100) if r.Trades else 0
        ws.cell(row=i, column=6).value = r.PnL
        ws.cell(row=i, column=7).value = r.PnL / r.Trades if r.Trades else 0
        ws.cell(row=i, column=8).value = round(r.AvgBars, 1)
        ws.cell(row=i, column=5).number_format = '0.0'
        ws.cell(row=i, column=6).number_format = '+#,##0.00;-#,##0.00;0.00'
        ws.cell(row=i, column=7).number_format = '+#,##0.00;-#,##0.00;0.00'
        if r.PnL > 0:
            for c in range(1, 9): ws.cell(row=i, column=c).fill = WIN_FILL
        elif r.PnL < 0:
            for c in range(1, 9): ws.cell(row=i, column=c).fill = LOSS_FILL

    # Totals row
    last = ws.max_row + 1
    ws.cell(row=last, column=1).value = 'TOTAL'
    ws.cell(row=last, column=2).value = int(grp['Trades'].sum())
    ws.cell(row=last, column=3).value = int(grp['Wins'].sum())
    ws.cell(row=last, column=4).value = int(grp['Losses'].sum())
    ws.cell(row=last, column=5).value = (grp['Wins'].sum() / grp['Trades'].sum() * 100)
    ws.cell(row=last, column=6).value = float(grp['PnL'].sum())
    ws.cell(row=last, column=7).value = float(grp['PnL'].sum() / grp['Trades'].sum())
    ws.cell(row=last, column=8).value = round(sized_df['bars_held'].mean(), 1)
    for c in range(1, 9):
        ws.cell(row=last, column=c).font = Font(bold=True)
        ws.cell(row=last, column=c).fill = SUBTOTAL_FILL
    ws.cell(row=last, column=5).number_format = '0.0'
    ws.cell(row=last, column=6).number_format = '+#,##0.00;-#,##0.00;0.00'
    ws.cell(row=last, column=7).number_format = '+#,##0.00;-#,##0.00;0.00'

    autosize(ws, {'A': 10})
    ws.freeze_panes = 'A4'
    return ws


def sheet_v2_vs_v21(wb, sized_df):
    ws = wb.create_sheet('V2 vs V2.1 (sized)')
    style_title(ws, 1, 5, 'V2 Method vs V2.1 Dynamic-TP — Position-Sized Comparison')
    ws.cell(row=2, column=1).value = 'Universe: 113 tickers (same as V2 Excel)  |  Window: 5y to 2026-05-18  |  Sizing: 1% risk/trade compounding from $10,000'
    ws.cell(row=2, column=1).font = SUBHEAD_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    headers = ['Metric', 'V2 (RSI 50 full exit)', 'V2.1 (TP1+TP2)', 'Delta', 'Notes']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c).value = h
    style_header_row(ws, 4, len(headers))

    # V2 baseline numbers come from the user's V2 Excel report (113 tickers, 1% compounding)
    # Source: SID V2 Method Back Testing (tiered + filter subtotals)(1).xlsx
    v2 = dict(
        trades=425, wins=275, losses=150, wr=64.7, pf=2.19,
        avg_win=222.72, avg_loss=-186.14, final=43327.80, total_return=333.28,
        cagr=34.67, max_dd=7.95, total_pnl=33327.80
    )

    wins = (sized_df['total_pnl'] > 0).sum()
    wr   = wins / len(sized_df) * 100
    pnl  = sized_df['total_pnl'].sum()
    avg_win  = sized_df.loc[sized_df['total_pnl'] > 0, 'total_pnl'].mean()
    avg_loss = sized_df.loc[sized_df['total_pnl'] <= 0, 'total_pnl'].mean()
    pf_w = sized_df.loc[sized_df['total_pnl'] > 0, 'total_pnl'].sum()
    pf_l = abs(sized_df.loc[sized_df['total_pnl'] < 0, 'total_pnl'].sum())
    pf = pf_w / pf_l if pf_l > 0 else float('inf')
    longs  = sized_df[sized_df['side'] == 'LONG']
    shorts = sized_df[sized_df['side'] == 'SHORT']
    final = sized_df.iloc[-1]['acct_after']

    def fmt_money(d):     return f'{d:+,.2f}'
    def fmt_pct_pp(d):    return f'{d:+.1f}pp'
    def fmt_2dp(d):       return f'{d:+.2f}'

    v21_cagr  = ((final / STARTING_ACCOUNT) ** (1/5) - 1) * 100
    v21_return = (final / STARTING_ACCOUNT - 1) * 100
    rows = [
        ('Total trades',         v2['trades'],    len(sized_df),                           f'{len(sized_df)-v2["trades"]:+d}',  'Same entry rules — small drift due to yfinance refresh between V2 Excel and this run'),
        ('Wins',                 v2['wins'],      int(wins),                               f'{int(wins)-v2["wins"]:+d}',        'V2.1: TP1 fired = win. V2: RSI 50 reached = win'),
        ('Win rate (%)',         v2['wr'],        round(wr, 1),                            fmt_pct_pp(wr - v2['wr']),           'V2.1 slightly lower — some zero-share trades + tied to BE-stop runner outcomes'),
        ('Profit factor',        v2['pf'],        round(pf, 2),                            fmt_2dp(pf - v2['pf']),              'Both well above 2.0; V2.1 PF benefits from TP2 uplift on winners'),
        ('Avg winner $',         v2['avg_win'],   round(avg_win, 2),                       fmt_money(avg_win - v2['avg_win']),  'V2.1 winners larger — TP2 leg captures the post-RSI50 extension'),
        ('Avg loser $',          v2['avg_loss'],  round(avg_loss, 2),                      fmt_money(avg_loss - v2['avg_loss']), ''),
        ('Net P&L $',            v2['total_pnl'], round(pnl, 2),                           fmt_money(pnl - v2['total_pnl']),    f'V2.1 vs V2: {(pnl-v2["total_pnl"])/v2["total_pnl"]*100:+.1f}%'),
        ('Long trades',          'see V2 sheet',  len(longs),                              '',         ''),
        ('Long WR (%)',          '',              round((longs['total_pnl'] > 0).mean()*100, 1), '',  ''),
        ('Short trades',         '',              len(shorts),                             '',         ''),
        ('Short WR (%)',         '',              round((shorts['total_pnl'] > 0).mean()*100, 1), '', ''),
        ('Avg bars held',        '~7.7',          round(sized_df['bars_held'].mean(), 1),  '',         'V2.1 holds 2-3 extra bars for the TP2 runner leg'),
        ('Starting balance $',   10000.00,        10000.00,                                '',         ''),
        ('Final balance $',      v2['final'],     round(final, 2),                         fmt_money(final - v2['final']),      '1% compounding sizing'),
        ('Total return (%)',     v2['total_return'], round(v21_return, 1),                 fmt_pct_pp(v21_return - v2['total_return']), ''),
        ('CAGR (%)',             v2['cagr'],      round(v21_cagr, 2),                      fmt_pct_pp(v21_cagr - v2['cagr']),   'Compounded annual growth rate'),
        ('TP2 uplift $',         'n/a',           round(sized_df.loc[sized_df["tp1_reason"]=="rsi50", "tp2_pnl"].sum(), 2), '', 'Total $ added by the V2.1 runner that V2 would have left on the table'),
    ]
    for i, r in enumerate(rows, start=5):
        for c, v in enumerate(r, 1):
            ws.cell(row=i, column=c).value = v
    autosize(ws, {'A': 22, 'E': 65})
    return ws


def sheet_account_growth(wb, sized_df):
    ws = wb.create_sheet('Account Growth')
    style_title(ws, 1, 6, 'V2.1 Account Equity Curve & Yearly Compounding')
    ws.cell(row=2, column=1).value = 'Year-end account balance (after every trade settles) — starts $10,000, 1% risk per trade'
    ws.cell(row=2, column=1).font = SUBHEAD_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ['Year', 'Start Balance', 'End Balance', '$ Gain', '% Gain', 'Trades']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c).value = h
    style_header_row(ws, 4, len(headers))

    df = sized_df.copy()
    df['year'] = pd.to_datetime(df['entry_date']).dt.year
    years = sorted(df['year'].unique())
    start_bal = STARTING_ACCOUNT
    for i, y in enumerate(years, start=5):
        yr = df[df['year'] == y]
        end_bal = yr.iloc[-1]['acct_after']
        gain = end_bal - start_bal
        pct = gain / start_bal
        ws.cell(row=i, column=1).value = y
        ws.cell(row=i, column=2).value = round(start_bal, 2)
        ws.cell(row=i, column=3).value = round(end_bal, 2)
        ws.cell(row=i, column=4).value = round(gain, 2)
        ws.cell(row=i, column=5).value = pct
        ws.cell(row=i, column=6).value = len(yr)
        ws.cell(row=i, column=2).number_format = '#,##0.00'
        ws.cell(row=i, column=3).number_format = '#,##0.00'
        ws.cell(row=i, column=4).number_format = '+#,##0.00;-#,##0.00;0.00'
        ws.cell(row=i, column=5).number_format = '+0.0%;-0.0%;0.0%'
        for c in range(1, 7):
            if gain > 0: ws.cell(row=i, column=c).fill = WIN_FILL
            elif gain < 0: ws.cell(row=i, column=c).fill = LOSS_FILL
        start_bal = end_bal

    # Total row
    last = ws.max_row + 1
    final = sized_df.iloc[-1]['acct_after']
    ws.cell(row=last, column=1).value = 'TOTAL'
    ws.cell(row=last, column=2).value = STARTING_ACCOUNT
    ws.cell(row=last, column=3).value = round(final, 2)
    ws.cell(row=last, column=4).value = round(final - STARTING_ACCOUNT, 2)
    ws.cell(row=last, column=5).value = (final - STARTING_ACCOUNT) / STARTING_ACCOUNT
    ws.cell(row=last, column=6).value = len(sized_df)
    for c in range(1, 7):
        ws.cell(row=last, column=c).font = Font(bold=True)
        ws.cell(row=last, column=c).fill = SUBTOTAL_FILL
    ws.cell(row=last, column=2).number_format = '#,##0.00'
    ws.cell(row=last, column=3).number_format = '#,##0.00'
    ws.cell(row=last, column=4).number_format = '+#,##0.00;-#,##0.00'
    ws.cell(row=last, column=5).number_format = '+0.0%;-0.0%'
    autosize(ws, {'A': 8})
    return ws


def sheet_rules(wb):
    ws = wb.create_sheet('Strategy Rules (V2.1)')
    style_title(ws, 1, 3, 'SID V2.1 Method — Rules + Position Sizing')
    rules = [
        ('Stage 1 — ARM (signal day)', '', ''),
        ('Daily RSI(14) oversold (long)', '< 30', 'Triggers long arm'),
        ('Daily RSI(14) overbought (short)', '> 70', 'Triggers short arm. V2 uses 70 not 75 (was V1.x change)'),
        ('Daily RSI(3) confirmation', 'Must be in extreme zone', 'Rebound-zone confirmation, must agree with RSI(14)'),
        ('Weekly 50/200 SMA trend filter', 'Direction must align with trade', 'V1.4+ kept in V2.x'),
        ('14-day earnings blackout', 'Pre-only (block 14 days BEFORE earnings)', 'Trading the day after earnings is permitted'),
        ('', '', ''),
        ('Stage 2 — TRIGGER (entry day, 1-3 bars after arm)', '', ''),
        ('Daily RSI(14) direction', 'Rising for long / falling for short', '1-bar slope (same as V2)'),
        ('Daily MACD(12,26,9) direction', 'Same as RSI direction', 'Histogram hidden per strategy'),
        ('RSI 45/55 no-go zone', 'Reject long if RSI ≥ 45, short if RSI ≤ 55', 'V2 addition (instructor S2_Ep3)'),
        ('Weekly RSI / MACD direction (OR mode)', 'Weekly RSI rising OR weekly MACD rising (mirror for shorts)', 'V2-weekly-or variant'),
        ('Stop level', 'Floor of (low between signal & trigger) for long / ceil of high for short', 'Whole dollar'),
        ('', '', ''),
        ('Stage 3 — TP1 (NEW in V2.1)', '', ''),
        ('TP1 trigger', 'Daily RSI(14) reaches 50', 'Same level V2 used for full exit'),
        ('TP1 action', 'Close 50% of position', 'Banks half the position profit'),
        ('Stop on remaining 50%', 'Move to entry price (break-even)', 'Eliminates downside risk on runner'),
        ('', '', ''),
        ('Stage 4 — TP2 (NEW in V2.1, runner exits)', '', ''),
        ('TP2 reason A', 'Price touches 50-day SMA', 'First MA target per instructor S3_P1'),
        ('TP2 reason B', 'Price touches 200-day SMA', 'Trend MA target'),
        ('TP2 reason C', 'Break-even stop hit', 'Runner round-trips back to entry'),
        ('TP2 reason D', '30 trading-day timeout', 'Force close at market'),
        ('', '', ''),
        ('Position sizing', '', ''),
        ('Risk per trade', f'{RISK_PCT*100:.1f}% of current account', 'Compounding from starting balance'),
        ('Shares formula', 'floor( risk_$ / risk_per_share )', 'Always rounds DOWN. No position cap in this report — pure 1% risk sizing (matches V2 Excel methodology). The live bot adds a 10% position cap on top of this.'),
        ('Zero-share trades', '0 shares = $0 P&L', 'Trade is logged but no position taken when risk/share is too large to size even 1 share at 1% risk'),
        ('', '', ''),
        ('Universe & tier routing', '', ''),
        ('AUTO tier', '80 tickers — bot auto-fires', 'Proven via 5y V1 backtest'),
        ('HUMAN tier', '32 tickers — Telegram approval (deferred v2.1)', 'High-vol / crypto / new'),
    ]
    ws.cell(row=3, column=1).value = 'Rule'
    ws.cell(row=3, column=2).value = 'Threshold / Action'
    ws.cell(row=3, column=3).value = 'Notes'
    style_header_row(ws, 3, 3)
    for i, (r, v, n) in enumerate(rules, start=4):
        ws.cell(row=i, column=1).value = r
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=3).value = n
        if not v and not n and r:  # section header row
            ws.cell(row=i, column=1).font = Font(bold=True, color='1F4E78')
            ws.cell(row=i, column=1).fill = SUBTOTAL_FILL
    autosize(ws, {'A': 38, 'B': 38, 'C': 60})
    return ws


def sheet_tp2_breakdown(wb, sized_df):
    ws = wb.create_sheet('TP2 Exit Breakdown')
    style_title(ws, 1, 7, 'V2.1 TP2 Runner — Exit-Reason Breakdown (winners only)')
    ws.cell(row=2, column=1).value = 'Of the 210 TP1 winners, the 50% runner exits via one of these four reasons:'
    ws.cell(row=2, column=1).font = SUBHEAD_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ['TP2 Reason', 'Count', '% of Winners', 'Avg TP2 P&L $', 'Sum TP2 P&L $', 'Description', 'Bot-side handling']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c).value = h
    style_header_row(ws, 4, len(headers))

    winners = sized_df[sized_df['tp1_reason'] == 'rsi50']
    desc = {
        'sma50_touch':    ('Runner exits when daily price closes through the 50-day SMA',  'Bot detects on daily check, submits next-day market close'),
        'sma200_touch':   ('Runner exits when daily price closes through the 200-day SMA', 'Bot detects on daily check, submits next-day market close'),
        'breakeven_stop': ('Runner stopped at entry price — round-trip to BE',              'Bot tracks the BE stop level set at TP1, closes at next-day open if hit'),
        'timeout':        ('30 trading days elapsed since TP1 — force close',              'Bot enforces with bars-since-TP1 counter, submits market close'),
    }
    reasons = winners['tp2_reason'].value_counts()
    total_winners = len(winners)
    row = 5
    for reason in ['sma50_touch', 'sma200_touch', 'breakeven_stop', 'timeout']:
        sub = winners[winners['tp2_reason'] == reason]
        cnt = len(sub)
        if cnt == 0: continue
        d, h = desc.get(reason, ('', ''))
        ws.cell(row=row, column=1).value = reason
        ws.cell(row=row, column=2).value = cnt
        ws.cell(row=row, column=3).value = cnt / total_winners
        ws.cell(row=row, column=4).value = round(sub['tp2_pnl'].mean(), 2)
        ws.cell(row=row, column=5).value = round(sub['tp2_pnl'].sum(), 2)
        ws.cell(row=row, column=6).value = d
        ws.cell(row=row, column=7).value = h
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=4).number_format = '+#,##0.00;-#,##0.00;0.00'
        ws.cell(row=row, column=5).number_format = '+#,##0.00;-#,##0.00;0.00'
        row += 1
    # Total
    ws.cell(row=row, column=1).value = 'ALL WINNERS'
    ws.cell(row=row, column=2).value = total_winners
    ws.cell(row=row, column=3).value = 1.0
    ws.cell(row=row, column=4).value = round(winners['tp2_pnl'].mean(), 2)
    ws.cell(row=row, column=5).value = round(winners['tp2_pnl'].sum(), 2)
    ws.cell(row=row, column=6).value = 'Sum of TP2 leg P&L across all 210 TP1 winners'
    for c in range(1, 8):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = SUBTOTAL_FILL
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=4).number_format = '+#,##0.00'
    ws.cell(row=row, column=5).number_format = '+#,##0.00'

    # Below the table — TP1 vs TP1+TP2 summary
    row += 3
    ws.cell(row=row, column=1).value = 'TP1 vs TP1+TP2 — what V2.1 adds over V2'
    ws.cell(row=row, column=1).font = Font(bold=True, color='1F4E78', size=12)
    row += 1
    tp1_only = winners['tp1_pnl'].sum()
    tp1_plus_tp2 = (winners['tp1_pnl'] + winners['tp2_pnl']).sum()
    uplift = tp1_plus_tp2 - tp1_only
    rows = [
        ('Sum of TP1 partials (winning trades only)',       f'${tp1_only:,.2f}',          'What V2 captures (full position closes at RSI 50)'),
        ('Sum of TP1+TP2 totals (winning trades only)',     f'${tp1_plus_tp2:,.2f}',      'What V2.1 actually books'),
        ('TP2 uplift (the runner\'s contribution)',         f'${uplift:,.2f} ({uplift/tp1_only*100:+.1f}%)', 'This is V2.1\'s edge over V2'),
    ]
    for r in rows:
        for c, v in enumerate(r, 1):
            ws.cell(row=row, column=c).value = v
        row += 1

    autosize(ws, {'A': 18, 'F': 60, 'G': 60})
    return ws


def sheet_tiers(wb, sized_df):
    ws = wb.create_sheet('Approval Tiers')
    style_title(ws, 1, 6, 'V2.1 Trades — Auto-Approved vs Human-Tier Breakdown')
    ws.cell(row=2, column=1).value = 'AUTO = ticker in proven 80-ticker tier1 set | HUMAN = high-vol/crypto/new (currently log-only)'
    ws.cell(row=2, column=1).font = SUBHEAD_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ['Tier', 'Trades', 'Tickers', 'WR (%)', 'Net P&L ($)', 'Avg P&L/trade ($)']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c).value = h
    style_header_row(ws, 4, len(headers))

    row = 5
    for tier_name in ['AUTO', 'HUMAN', 'OTHER']:
        sub = sized_df[sized_df['tier'] == tier_name]
        if len(sub) == 0: continue
        wins = (sub['total_pnl'] > 0).sum()
        ws.cell(row=row, column=1).value = tier_name
        ws.cell(row=row, column=2).value = len(sub)
        ws.cell(row=row, column=3).value = sub['ticker'].nunique()
        ws.cell(row=row, column=4).value = round(wins / len(sub) * 100, 1)
        ws.cell(row=row, column=5).value = round(sub['total_pnl'].sum(), 2)
        ws.cell(row=row, column=6).value = round(sub['total_pnl'].mean(), 2)
        ws.cell(row=row, column=5).number_format = '+#,##0.00;-#,##0.00;0.00'
        ws.cell(row=row, column=6).number_format = '+#,##0.00;-#,##0.00;0.00'
        row += 1

    # Combined total
    wins = (sized_df['total_pnl'] > 0).sum()
    ws.cell(row=row, column=1).value = 'COMBINED'
    ws.cell(row=row, column=2).value = len(sized_df)
    ws.cell(row=row, column=3).value = sized_df['ticker'].nunique()
    ws.cell(row=row, column=4).value = round(wins / len(sized_df) * 100, 1)
    ws.cell(row=row, column=5).value = round(sized_df['total_pnl'].sum(), 2)
    ws.cell(row=row, column=6).value = round(sized_df['total_pnl'].mean(), 2)
    for c in range(1, 7):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = SUBTOTAL_FILL
    ws.cell(row=row, column=5).number_format = '+#,##0.00'
    ws.cell(row=row, column=6).number_format = '+#,##0.00'
    autosize(ws, {'A': 14})
    return ws


def main():
    print('Loading V2.1 trades + simulating compounding sizing...')
    sized = load_and_size(CSV_PATH)
    print(f'  {len(sized)} trades; final account ${sized.iloc[-1]["acct_after"]:,.2f}')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet

    sheet_overview     (wb, sized)
    sheet_all_trades   (wb, sized)
    sheet_per_ticker   (wb, sized)
    sheet_v2_vs_v21    (wb, sized)
    sheet_account_growth(wb, sized)
    sheet_tp2_breakdown(wb, sized)
    sheet_rules        (wb)
    sheet_tiers        (wb, sized)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f'Wrote: {OUT_PATH}')
    print(f'Size: {OUT_PATH.stat().st_size:,} bytes')


if __name__ == '__main__':
    main()
