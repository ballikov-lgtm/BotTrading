"""
build-v2-xlsx.py — Build "SID V2 Method Back Testing.xlsx" from
backtest-v2-validation-report.json, with INSTRUCTOR-STYLE POSITION SIZING.

Instructor's method (from S3_P3 Position Sizing + S3_Ep4 Risk Management):
  Risk $ = Account × Risk %         (1% default per instructor)
  Risk per share = |Entry − Stop|
  Shares = ROUNDDOWN(Risk $ / Risk per share)
  Position $ = Shares × Entry
  P&L = Shares × (Exit − Entry) [long] or × (Entry − Exit) [short]

Position size scales with account balance. Each trade has its own risk
amount, shares, position size, and P&L — compounding chronologically.

Default: $10K start, 1% risk per trade.

Sheets:
  1. All Trades (V2)             — sized trades with running account balance
  2. Per-Ticker Summary          — recomputed with sized P&L
  3. V1 vs V2 Comparison         — aggregate side-by-side (fixed $200 vs sized)
  4. Account Growth              — equity curve, year-by-year balances
  5. Strategy Rules (V2)         — updated with 1% risk + sizing methodology
"""
import json
import sys
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

JSON_PATH = Path(__file__).parent.parent / 'backtest-v2-validation-report.json'
OUT_PATH  = Path(__file__).parent.parent / 'SID V2 Method Back Testing.xlsx'

STARTING_BALANCE = 10000.0
RISK_PCT         = 0.01      # 1% per instructor default (S3_Ep4)

# Auto-approved tickers — proven via 5y V1 backtest, refined-47 + tier1 expansion
# (the original 80-ticker universe). Anything OUTSIDE this list requires
# human approval via Telegram (see V2-TELEGRAM-APPROVAL-SPEC.md).
AUTO_APPROVED = {
    'AAPL','ABBV','ABT','ADBE','AMAT','AMD','AMZN','AVGO','AXP','B',
    'BA','BLK','CAT','COST','CRM','CVX','DE','DIA','DIS','EEM',
    'EFA','F','GDX','GE','GLD','GOOG','GS','HD','HON','IBB',
    'IBM','INTC','IWM','IYR','JNJ','JPM','KHC','LLY','LMT','LRCX',
    'MCD','MDLZ','META','MRK','NKE','NOW','NUGT','NVDA','ORCL','PFE',
    'PYPL','QQQ','RIOT','RTX','SBUX','SCHW','SLV','SPY','SQQQ','TGT',
    'TNA','TQQQ','TSLA','TZA','UNH','V','WFC','WMT','XHB','XLC',
    'XLE','XLF','XLI','XLK','XLP','XLRE','XLU','XLV','XLY','XOM',
}

if not JSON_PATH.exists():
    print(f'ERROR: {JSON_PATH} not found. Run backtest-sid-v2.py first.')
    sys.exit(1)

report = json.loads(JSON_PATH.read_text(encoding='utf-8'))

trades_by_variant = report['all_trades']
agg = report['aggregates']

# Headline V2 variant priority
V2_VARIANT_PRIORITY = ['v2-weekly-or', 'v2-slope', 'v2-method', 'v2-nogo-only']
headline_variant = None
for cand in V2_VARIANT_PRIORITY:
    if cand in trades_by_variant and len(trades_by_variant[cand]) >= 100:
        headline_variant = cand
        break
if headline_variant is None:
    non_v1 = {k: v for k, v in trades_by_variant.items() if k != 'v1.7-shipped'}
    if non_v1:
        headline_variant = max(non_v1, key=lambda k: len(non_v1[k]))
    else:
        headline_variant = 'v2-method'

v2_trades_raw = trades_by_variant.get(headline_variant, [])
v1_trades_raw = trades_by_variant.get('v1.7-shipped', [])
v2_agg_picked = agg.get(headline_variant, {})

# ─── Position-sized P&L computation (instructor methodology) ────────────────
def compute_position_sized(trades_in, starting_balance, risk_pct):
    """Process trades chronologically, scaling position size with running
    account balance per instructor's S3_P3 method. Returns enriched trade
    list + summary stats.

    Approximation: treats trades as sequential (P&L realised before next
    trade's entry). For 0-3 concurrent positions this is small error.
    """
    trades = sorted(trades_in, key=lambda t: (t['entry_date'], t['ticker']))
    sized = []
    balance = starting_balance
    peak = starting_balance
    max_dd = 0.0

    for t in trades:
        balance_before = balance
        risk_dollars = balance_before * risk_pct
        risk_per_share = abs(float(t['entry_price']) - float(t['stop']))

        if risk_per_share <= 0 or t['entry_price'] <= 0:
            shares = 0
        else:
            shares = int(risk_dollars / risk_per_share)
            # Cash trading — no margin: cap shares so position fits in account
            max_shares_cash = int(balance_before / float(t['entry_price']))
            shares = min(shares, max_shares_cash)
            if shares < 1:
                shares = 0  # Trade skipped — risk too small for one share

        position_value = shares * float(t['entry_price'])

        if t['side'] == 'LONG':
            pnl_sized = (float(t['exit_price']) - float(t['entry_price'])) * shares
        else:
            pnl_sized = (float(t['entry_price']) - float(t['exit_price'])) * shares

        balance_after = balance_before + pnl_sized

        # Track drawdown
        if balance_after > peak:
            peak = balance_after
        dd = (peak - balance_after) / peak * 100 if peak > 0 else 0
        if dd > max_dd:
            max_dd = dd

        s = dict(t)
        s.update({
            'balance_before': round(balance_before, 2),
            'risk_pct': risk_pct,
            'risk_dollars': round(risk_dollars, 2),
            'risk_per_share': round(risk_per_share, 4),
            'shares_sized': shares,
            'position_value': round(position_value, 2),
            'pnl_sized': round(pnl_sized, 2),
            'balance_after': round(balance_after, 2),
            'pnl_pct_sized': round((pnl_sized / balance_before) * 100, 3) if balance_before > 0 else 0,
        })
        sized.append(s)
        balance = balance_after

    # Summary stats
    final_balance = balance
    total_return_pct = ((final_balance - starting_balance) / starting_balance) * 100
    # CAGR
    if sized:
        start_dt = datetime.fromisoformat(sized[0]['entry_date'])
        end_dt   = datetime.fromisoformat(sized[-1]['exit_date'])
        years = (end_dt - start_dt).days / 365.25
        cagr = ((final_balance / starting_balance) ** (1/years) - 1) * 100 if years > 0 else 0
    else:
        years = 0; cagr = 0
    wins = [s for s in sized if s['pnl_sized'] > 0]
    losses = [s for s in sized if s['pnl_sized'] <= 0]
    summary = {
        'starting_balance': starting_balance,
        'final_balance': round(final_balance, 2),
        'total_return_pct': round(total_return_pct, 2),
        'cagr_pct': round(cagr, 2),
        'years': round(years, 2),
        'total_trades': len(sized),
        'wins': len(wins),
        'losses': len(losses),
        'win_rate': round(len(wins) / len(sized) * 100, 1) if sized else 0,
        'avg_win_dollar': round(sum(s['pnl_sized'] for s in wins) / len(wins), 2) if wins else 0,
        'avg_loss_dollar': round(sum(s['pnl_sized'] for s in losses) / len(losses), 2) if losses else 0,
        'max_drawdown_pct': round(max_dd, 2),
        'profit_factor': round(
            abs(sum(s['pnl_sized'] for s in wins) /
                sum(s['pnl_sized'] for s in losses)), 2
        ) if losses and sum(s['pnl_sized'] for s in losses) != 0 else 0,
    }
    return sized, summary

v2_sorted, v2_summary = compute_position_sized(v2_trades_raw, STARTING_BALANCE, RISK_PCT)
_, v1_summary = compute_position_sized(v1_trades_raw, STARTING_BALANCE, RISK_PCT)

print(f'Headline V2 variant: {headline_variant}  ({len(v2_trades_raw)} trades)')
print(f'V2 sized — start ${STARTING_BALANCE:,.0f}, end ${v2_summary["final_balance"]:,.2f}, '
      f'return {v2_summary["total_return_pct"]}%, CAGR {v2_summary["cagr_pct"]}%, '
      f'max DD {v2_summary["max_drawdown_pct"]}%')
print(f'V1 sized — start ${STARTING_BALANCE:,.0f}, end ${v1_summary["final_balance"]:,.2f}, '
      f'return {v1_summary["total_return_pct"]}%, CAGR {v1_summary["cagr_pct"]}%, '
      f'max DD {v1_summary["max_drawdown_pct"]}%')

# ─── Theme ──────────────────────────────────────────────────────────────
THEME = {
    'title_bg':       '1A365D',  # Dark navy
    'title_fg':       'FFFFFF',
    'header_bg':      '2C5282',
    'header_fg':      'FFFFFF',
    'subheader_bg':   '4299E1',
    'subheader_fg':   'FFFFFF',
    'row_alt':        'EBF8FF',
    'row_white':      'FFFFFF',
    'win_bg':         'C6F6D5',
    'win_fg':         '22543D',
    'loss_bg':        'FED7D7',
    'loss_fg':        '742A2A',
    'subtotal_bg':    'FEF3C7',
    'subtotal_fg':    '744210',
    'border':         'CBD5E0',
    'long_bg':        'BEE3F8',
    'short_bg':       'FBB6CE',
    'cyan_bg':        'E6FFFA',
    'magenta_bg':     'FFE4F1',
    'sizing_bg':      'FFFAF0',  # Soft cream for position-sizing columns
}

THIN = Side(border_style='thin', color=THEME['border'])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ARIAL = 'Arial'

wb = Workbook()

# ============================================================================
# Sheet 1: All Trades (V2) — position-sized
# ============================================================================
ws = wb.active
ws.title = 'All Trades (V2 sized)'

TITLE = (f"SID V2 Method — Position-Sized Backtest "
         f"(${STARTING_BALANCE:,.0f} start, {RISK_PCT*100:.1f}% risk/trade, "
         f"{report['backtest_years']}y, {len(report['tickers'])} tickers)")
ws.merge_cells('A1:Y1')
ws['A1'] = TITLE
ws['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

subtitle = (
    f"{v2_summary['total_trades']} trades  |  WR {v2_summary['win_rate']}%  |  "
    f"PF {v2_summary['profit_factor']}  |  "
    f"Start ${STARTING_BALANCE:,.0f} → End ${v2_summary['final_balance']:,.2f}  "
    f"({v2_summary['total_return_pct']:+.1f}% total, {v2_summary['cagr_pct']:.1f}% CAGR over {v2_summary['years']}y)  |  "
    f"Max DD {v2_summary['max_drawdown_pct']}%  |  Generated {datetime.now().strftime('%Y-%m-%d')}"
)
ws.merge_cells('A2:Y2')
ws['A2'] = subtitle
ws['A2'].font = Font(name=ARIAL, size=11, color='2D3748', italic=True)
ws['A2'].fill = PatternFill('solid', fgColor='F7FAFC')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

HEADER_ROW = 3
# kind values: int|text|side|date|money|risk_pct|pct|rsi|flag|reason|outcome
COLS = [
    ('Trade #',          'trade_no',          'int',     7,  False),
    ('Ticker',           'ticker',            'text',    8,  False),
    ('Side',             'side',              'side',    8,  False),
    ('Entry Date',       'entry_date',        'date',    12, False),
    ('Entry $',          'entry_price',       'money',   11, False),
    ('Stop $',           'stop',              'money',   10, False),
    ('Exit Date',        'exit_date',         'date',    12, False),
    ('Exit $',           'exit_price',        'money',   11, False),
    ('Exit Reason',      'exit_reason',       'reason',  11, False),
    # Position sizing block (highlighted)
    ('Acct Before',      'balance_before',    'money',   13, True),
    ('Risk %',           'risk_pct',          'risk_pct',9,  True),
    ('Risk $',           'risk_dollars',      'money',   11, True),
    ('Risk/Share',       'risk_per_share',    'money',   11, True),
    ('Position $',       'position_value',    'money',   13, True),
    ('Shares',           'shares_sized',      'int',     9,  True),
    # Outcome block
    ('P&L $',            'pnl_sized',         'pnl',     12, False),
    ('P&L %',            'pnl_pct_sized',     'pct_raw', 10, False),
    ('Acct After',       'balance_after',     'money',   13, True),
    ('Bars Held',        'bars_held',         'int',     9,  False),
    ('RSI @ Entry',      'entry_rsi',         'rsi',     11, False),
    ('Wkly RSI',         'wk_rsi_rising_at_entry',  'flag', 10, False),
    ('Wkly MACD',        'wk_macd_rising_at_entry', 'flag', 10, False),
    ('Approval',         None,                'approval', 11, False),
    ('Outcome',          None,                'outcome', 10, False),
]

for col_idx, (label, _, _, width, _) in enumerate(COLS, 1):
    c = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[HEADER_ROW].height = 34

DATA_START = HEADER_ROW + 1
for i, t in enumerate(v2_sorted):
    row = DATA_START + i
    is_win = float(t['pnl_sized']) > 0
    side = t['side']

    for col_idx, (_, key, kind, _, is_sizing) in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col_idx)
        # Pick value
        if key is None and kind == 'approval':
            tier = 'AUTO' if t['ticker'] in AUTO_APPROVED else 'HUMAN'
            c.value = tier
        elif key is None and kind == 'outcome':
            c.value = 'WIN' if is_win else 'LOSS'
        elif key == 'trade_no':
            c.value = i + 1; c.number_format = '0'
        elif kind == 'int':
            c.value = int(t.get(key) or 0); c.number_format = '0'
        elif kind == 'money':
            c.value = float(t.get(key) or 0); c.number_format = '$#,##0.00;[Red]-$#,##0.00'
        elif kind == 'pnl':
            c.value = float(t.get(key) or 0); c.number_format = '$#,##0.00;[Red]-$#,##0.00'
        elif kind == 'risk_pct':
            c.value = float(t.get(key) or 0); c.number_format = '0.0%'
        elif kind == 'pct_raw':
            c.value = float(t.get(key) or 0) / 100; c.number_format = '0.00%;[Red]-0.00%'
        elif kind == 'rsi':
            c.value = float(t.get(key) or 0); c.number_format = '0.0'
        elif kind == 'flag':
            v = t.get(key)
            c.value = 'YES' if v else 'NO' if v is False else '—'
        elif kind == 'date':
            c.value = t.get(key); c.number_format = '@'
        else:
            c.value = t.get(key)

        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Row stripe (alt OR sizing-column tint)
        if is_sizing:
            c.fill = PatternFill('solid', fgColor=THEME['sizing_bg'])
        else:
            c.fill = PatternFill('solid', fgColor=(THEME['row_white'] if i % 2 == 0 else THEME['row_alt']))

        if kind == 'side':
            if side == 'LONG':
                c.fill = PatternFill('solid', fgColor=THEME['long_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color='2C5282')
            else:
                c.fill = PatternFill('solid', fgColor=THEME['short_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color='9B2C2C')

        if kind == 'flag':
            v = t.get(key)
            if v is True:
                c.fill = PatternFill('solid', fgColor=THEME['win_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
            elif v is False:
                c.fill = PatternFill('solid', fgColor='E2E8F0')
                c.font = Font(name=ARIAL, size=10, color='4A5568')

        if kind == 'approval':
            tier = 'AUTO' if t['ticker'] in AUTO_APPROVED else 'HUMAN'
            if tier == 'AUTO':
                c.fill = PatternFill('solid', fgColor=THEME['win_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
            else:
                c.fill = PatternFill('solid', fgColor='FED7AA')  # Amber/warning
                c.font = Font(name=ARIAL, size=10, bold=True, color='9C4221')

        if kind == 'outcome':
            if is_win:
                c.fill = PatternFill('solid', fgColor=THEME['win_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
            else:
                c.fill = PatternFill('solid', fgColor=THEME['loss_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['loss_fg'])

        if kind == 'pnl':
            if is_win:
                c.font = Font(name=ARIAL, size=10, bold=True, color='22543D')
            else:
                c.font = Font(name=ARIAL, size=10, bold=True, color='742A2A')

DATA_END = DATA_START + len(v2_sorted) - 1
ws.freeze_panes = 'D4'
if DATA_END >= DATA_START:
    ws.auto_filter.ref = f"A{HEADER_ROW}:Y{DATA_END}"

# ── Filter-aware subtotal row ───────────────────────────────────────────────
# Uses SUBTOTAL(function, range) for live recalculation as filters toggle.
# Win Rate column uses SUMPRODUCT+SUBTOTAL(3, OFFSET(...)) to count visible
# "WIN" outcomes vs total visible rows — recalculates as filters change.
SUBTOTAL_ROW = DATA_END + 1
SUBTOTAL_LABELS = {
    1:  ('subtotal_label', None),                          # Trade # — label
    2:  ('count_text',     None),                          # Ticker — "N trades"
    3:  ('blank',          None),                          # Side
    4:  ('blank',          None),                          # Entry Date
    5:  ('avg_money',      'E'),                           # Entry $
    6:  ('avg_money',      'F'),                           # Stop $
    7:  ('blank',          None),                          # Exit Date
    8:  ('avg_money',      'H'),                           # Exit $
    9:  ('blank',          None),                          # Exit Reason
    10: ('blank',          None),                          # Acct Before (not meaningful aggregated)
    11: ('blank',          None),                          # Risk %
    12: ('sum_money',      'L'),                           # Risk $
    13: ('avg_money',      'M'),                           # Risk/Share
    14: ('sum_money',      'N'),                           # Position $
    15: ('sum_int',        'O'),                           # Shares
    16: ('sum_money',      'P'),                           # P&L $
    17: ('avg_pct_raw',    'Q'),                           # P&L %
    18: ('blank',          None),                          # Acct After
    19: ('avg_decimal',    'S'),                           # Bars Held
    20: ('avg_decimal',    'T'),                           # RSI @ Entry
    21: ('blank',          None),                          # Wkly RSI
    22: ('blank',          None),                          # Wkly MACD
    23: ('blank',          None),                          # Approval
    24: ('win_rate',       'X'),                           # Outcome → WIN RATE %
}

def get_col_letter_from_idx(idx):
    return get_column_letter(idx)

for col_idx, (kind, col_ref) in SUBTOTAL_LABELS.items():
    c = ws.cell(row=SUBTOTAL_ROW, column=col_idx)
    rng_start = f"{col_ref}{DATA_START}" if col_ref else None
    rng_end   = f"{col_ref}{DATA_END}"   if col_ref else None

    if kind == 'subtotal_label':
        c.value = 'SUBTOTAL ▼'
    elif kind == 'count_text':
        # Number of VISIBLE trades after filtering
        c.value = f'=SUBTOTAL(3,B{DATA_START}:B{DATA_END})&" trades"'
    elif kind == 'sum_money':
        c.value = f'=SUBTOTAL(9,{rng_start}:{rng_end})'
        c.number_format = '$#,##0.00;[Red]-$#,##0.00'
    elif kind == 'sum_int':
        c.value = f'=SUBTOTAL(9,{rng_start}:{rng_end})'
        c.number_format = '#,##0'
    elif kind == 'avg_money':
        c.value = f'=IFERROR(SUBTOTAL(1,{rng_start}:{rng_end}),0)'
        c.number_format = '$#,##0.00'
    elif kind == 'avg_pct_raw':
        c.value = f'=IFERROR(SUBTOTAL(1,{rng_start}:{rng_end}),0)'
        c.number_format = '0.00%;[Red]-0.00%'
    elif kind == 'avg_decimal':
        c.value = f'=IFERROR(SUBTOTAL(1,{rng_start}:{rng_end}),0)'
        c.number_format = '0.0'
    elif kind == 'win_rate':
        # Counts visible "WIN" outcomes / total visible rows
        # SUMPRODUCT iterates each row; SUBTOTAL(3, OFFSET(...)) returns
        # 1 if visible, 0 if hidden by filter. Multiplied by ="WIN" gives
        # count of visible wins. Divided by total visible (with MAX(...,1)
        # to avoid div-by-zero when everything filtered out).
        c.value = (
            f'=IFERROR(SUMPRODUCT(SUBTOTAL(3,OFFSET({col_ref}{DATA_START},'
            f'ROW({col_ref}{DATA_START}:{col_ref}{DATA_END})-ROW({col_ref}{DATA_START}),0))'
            f'*({col_ref}{DATA_START}:{col_ref}{DATA_END}="WIN"))'
            f'/MAX(SUBTOTAL(3,{col_ref}{DATA_START}:{col_ref}{DATA_END}),1),0)'
        )
        c.number_format = '0.0%'
    # blank → leave value as None

    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['subtotal_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['subtotal_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER

ws.row_dimensions[SUBTOTAL_ROW].height = 28

# ============================================================================
# Sheet 2: Per-Ticker Summary
# ============================================================================
ws2 = wb.create_sheet('Per-Ticker Summary')
ws2.merge_cells('A1:G1')
ws2['A1'] = f'V2 Method (sized at {RISK_PCT*100:.1f}% risk) — Per-Ticker Performance'
ws2['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws2['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 32

by_ticker = {}
for t in v2_sorted:
    tk = t['ticker']
    by_ticker.setdefault(tk, []).append(t)

ticker_rows = []
for tk, ts in by_ticker.items():
    n = len(ts)
    w = sum(1 for t in ts if float(t['pnl_sized']) > 0)
    pnl = sum(float(t['pnl_sized']) for t in ts)
    avg_bars = sum(int(t['bars_held']) for t in ts) / n
    ticker_rows.append((tk, n, w, n - w, w / n * 100 if n else 0, pnl, avg_bars))
ticker_rows.sort(key=lambda r: -r[5])

T_HEADER_ROW = 3
T_HEADERS = [('Ticker',12),('Trades',10),('Wins',9),('Losses',10),
             ('Win Rate (%)',14),('Net P&L ($)',14),('Avg Bars Held',14)]
for col_idx, (label, width) in enumerate(T_HEADERS, 1):
    c = ws2.cell(row=T_HEADER_ROW, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[T_HEADER_ROW].height = 32

T_DATA_START = T_HEADER_ROW + 1
for i, row in enumerate(ticker_rows):
    r = T_DATA_START + i
    tk, n, w, l, wr, pnl, avg_bars = row
    for col_idx, val in enumerate(row, 1):
        c = ws2.cell(row=r, column=col_idx, value=val)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        c.fill = PatternFill('solid', fgColor=(THEME['row_white'] if i % 2 == 0 else THEME['row_alt']))
    ws2.cell(row=r, column=2).number_format = '0'
    ws2.cell(row=r, column=3).number_format = '0'
    ws2.cell(row=r, column=4).number_format = '0'
    ws2.cell(row=r, column=5).number_format = '0.0'
    ws2.cell(row=r, column=6).number_format = '$#,##0.00;[Red]-$#,##0.00'
    ws2.cell(row=r, column=7).number_format = '0.0'

    c_pnl = ws2.cell(row=r, column=6)
    if pnl > 0:
        c_pnl.font = Font(name=ARIAL, size=10, bold=True, color='22543D')
    elif pnl < 0:
        c_pnl.font = Font(name=ARIAL, size=10, bold=True, color='742A2A')
    c_wr = ws2.cell(row=r, column=5)
    if wr >= 65:
        c_wr.fill = PatternFill('solid', fgColor=THEME['win_bg'])
        c_wr.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
    elif wr < 40:
        c_wr.fill = PatternFill('solid', fgColor=THEME['loss_bg'])
        c_wr.font = Font(name=ARIAL, size=10, bold=True, color=THEME['loss_fg'])

if ticker_rows:
    T_TOTAL_ROW = T_DATA_START + len(ticker_rows)
    ws2.cell(row=T_TOTAL_ROW, column=1).value = 'TOTAL'
    ws2.cell(row=T_TOTAL_ROW, column=2).value = f'=SUM(B{T_DATA_START}:B{T_TOTAL_ROW-1})'
    ws2.cell(row=T_TOTAL_ROW, column=3).value = f'=SUM(C{T_DATA_START}:C{T_TOTAL_ROW-1})'
    ws2.cell(row=T_TOTAL_ROW, column=4).value = f'=SUM(D{T_DATA_START}:D{T_TOTAL_ROW-1})'
    ws2.cell(row=T_TOTAL_ROW, column=5).value = f'=C{T_TOTAL_ROW}/B{T_TOTAL_ROW}*100'
    ws2.cell(row=T_TOTAL_ROW, column=6).value = f'=SUM(F{T_DATA_START}:F{T_TOTAL_ROW-1})'
    ws2.cell(row=T_TOTAL_ROW, column=7).value = f'=AVERAGE(G{T_DATA_START}:G{T_TOTAL_ROW-1})'
    for col_idx in range(1, 8):
        c = ws2.cell(row=T_TOTAL_ROW, column=col_idx)
        c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['subtotal_fg'])
        c.fill = PatternFill('solid', fgColor=THEME['subtotal_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws2.cell(row=T_TOTAL_ROW, column=2).number_format = '0'
    ws2.cell(row=T_TOTAL_ROW, column=3).number_format = '0'
    ws2.cell(row=T_TOTAL_ROW, column=4).number_format = '0'
    ws2.cell(row=T_TOTAL_ROW, column=5).number_format = '0.0'
    ws2.cell(row=T_TOTAL_ROW, column=6).number_format = '$#,##0.00;[Red]-$#,##0.00'
    ws2.cell(row=T_TOTAL_ROW, column=7).number_format = '0.0'

ws2.freeze_panes = 'A4'

# ============================================================================
# Sheet 3: V1 vs V2 Comparison (sized)
# ============================================================================
ws3 = wb.create_sheet('V1 vs V2 (sized)')
ws3.merge_cells('A1:E1')
ws3['A1'] = f'V1 Baseline vs V2 Method — Both Position-Sized at {RISK_PCT*100:.1f}% Risk'
ws3['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws3['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 32

ws3.merge_cells('A2:E2')
ws3['A2'] = (f"Universe: {len(report['tickers'])} tickers  |  "
             f"Window: {report['backtest_start']} → {report['backtest_end']}  ({report['backtest_years']}y)  |  "
             f"Starting balance: ${STARTING_BALANCE:,.0f}")
ws3['A2'].font = Font(name=ARIAL, size=11, italic=True, color='2D3748')
ws3['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws3['A2'].fill = PatternFill('solid', fgColor='F7FAFC')
ws3.row_dimensions[2].height = 22

C_HEADER_ROW = 4
HEADERS = ['Metric', 'V1.7-shipped (baseline)', f'V2 ({headline_variant})', 'Delta', 'Notes']
for col_idx, label in enumerate(HEADERS, 1):
    c = ws3.cell(row=C_HEADER_ROW, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
ws3.row_dimensions[C_HEADER_ROW].height = 32

def fmt_delta_num(a, b, suffix=''):
    d = b - a
    return f"{d:+,.2f}{suffix}" if isinstance(d, float) else f"{d:+,}{suffix}"

rows = [
    ('Total trades',     v1_summary['total_trades'],     v2_summary['total_trades'],     fmt_delta_num(v1_summary['total_trades'], v2_summary['total_trades']), 'V2 = stricter, fewer trades'),
    ('Wins',             v1_summary['wins'],             v2_summary['wins'],             fmt_delta_num(v1_summary['wins'], v2_summary['wins']), ''),
    ('Losses',           v1_summary['losses'],           v2_summary['losses'],           fmt_delta_num(v1_summary['losses'], v2_summary['losses']), ''),
    ('Win Rate (%)',     v1_summary['win_rate'],         v2_summary['win_rate'],         f"{v2_summary['win_rate'] - v1_summary['win_rate']:+.1f}pp", 'Community claim band: 65-75%'),
    ('Profit Factor',    v1_summary['profit_factor'],    v2_summary['profit_factor'],    f"{v2_summary['profit_factor'] - v1_summary['profit_factor']:+.2f}", 'Higher = wins outweigh losses more'),
    ('Starting balance', f"${v1_summary['starting_balance']:,.0f}", f"${v2_summary['starting_balance']:,.0f}", '', 'Both start at same balance'),
    ('Final balance',    f"${v1_summary['final_balance']:,.2f}",    f"${v2_summary['final_balance']:,.2f}",    f"${v2_summary['final_balance'] - v1_summary['final_balance']:+,.2f}", 'After all trades, compounded'),
    ('Total return (%)', f"{v1_summary['total_return_pct']}%",      f"{v2_summary['total_return_pct']}%",      f"{v2_summary['total_return_pct'] - v1_summary['total_return_pct']:+.1f}pp", ''),
    ('CAGR (%)',         f"{v1_summary['cagr_pct']}%",              f"{v2_summary['cagr_pct']}%",              f"{v2_summary['cagr_pct'] - v1_summary['cagr_pct']:+.1f}pp", 'Compounded annual growth rate'),
    ('Max drawdown (%)', f"{v1_summary['max_drawdown_pct']}%",      f"{v2_summary['max_drawdown_pct']}%",      f"{v2_summary['max_drawdown_pct'] - v1_summary['max_drawdown_pct']:+.1f}pp", 'Peak-to-trough decline'),
    ('Avg win ($)',      f"${v1_summary['avg_win_dollar']:.2f}",    f"${v2_summary['avg_win_dollar']:.2f}",    f"${v2_summary['avg_win_dollar'] - v1_summary['avg_win_dollar']:+.2f}", 'Scales with account growth'),
    ('Avg loss ($)',     f"${v1_summary['avg_loss_dollar']:.2f}",   f"${v2_summary['avg_loss_dollar']:.2f}",   f"${v2_summary['avg_loss_dollar'] - v1_summary['avg_loss_dollar']:+.2f}", 'Scales with account growth'),
]
C_DATA_START = C_HEADER_ROW + 1
for i, r_data in enumerate(rows):
    r = C_DATA_START + i
    for col_idx, val in enumerate(r_data, 1):
        c = ws3.cell(row=r, column=col_idx, value=val)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        c.fill = PatternFill('solid', fgColor=(THEME['row_white'] if i % 2 == 0 else THEME['row_alt']))
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws3.cell(row=r, column=1).font = Font(name=ARIAL, size=10, bold=True, color='2D3748')

    # Highlight key rows
    if r_data[0] in ('Win Rate (%)', 'CAGR (%)', 'Final balance', 'Max drawdown (%)'):
        v1_val = r_data[1]
        v2_val = r_data[2]
        # extract numeric for comparison
        def _num(x):
            try: return float(str(x).replace('$','').replace(',','').replace('%','').replace('pp','').strip())
            except: return 0
        v1n = _num(v1_val); v2n = _num(v2_val)
        # For drawdown, lower is better; for others higher is better
        cell = ws3.cell(row=r, column=3)
        if r_data[0] == 'Max drawdown (%)':
            better = v2n < v1n
        else:
            better = v2n > v1n
        if better:
            cell.fill = PatternFill('solid', fgColor=THEME['win_bg'])
            cell.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
        else:
            cell.fill = PatternFill('solid', fgColor=THEME['loss_bg'])
            cell.font = Font(name=ARIAL, size=10, bold=True, color=THEME['loss_fg'])

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 24
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 42

# ============================================================================
# Sheet 4: Account Growth (equity curve + year summaries)
# ============================================================================
ws4 = wb.create_sheet('Account Growth')
ws4.merge_cells('A1:F1')
ws4['A1'] = 'V2 Account Equity Curve & Year-by-Year Summary'
ws4['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws4['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 32

# Year-by-year section
ws4['A3'] = 'Year-end account balance (after each calendar year of trades)'
ws4['A3'].font = Font(name=ARIAL, size=12, bold=True, color='2D3748')
ws4.merge_cells('A3:F3')

year_headers = ['Year', 'Start Balance', 'End Balance', '$ Gain', '% Gain', 'Trades']
for col_idx, label in enumerate(year_headers, 1):
    c = ws4.cell(row=5, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER
    ws4.column_dimensions[get_column_letter(col_idx)].width = 16

# Compute year-end balances
year_data = {}
for t in v2_sorted:
    y = int(t['exit_date'][:4])
    year_data.setdefault(y, []).append(t)

prev_end = STARTING_BALANCE
row_idx = 6
for y in sorted(year_data.keys()):
    trades_y = year_data[y]
    start_bal = prev_end
    # End balance = last trade's balance_after in that year
    end_bal = trades_y[-1]['balance_after']
    gain = end_bal - start_bal
    gain_pct = (gain / start_bal) * 100 if start_bal else 0

    row = (y, start_bal, end_bal, gain, gain_pct / 100, len(trades_y))
    for col_idx, val in enumerate(row, 1):
        c = ws4.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        c.fill = PatternFill('solid', fgColor=(THEME['row_white'] if row_idx % 2 == 0 else THEME['row_alt']))
    ws4.cell(row=row_idx, column=1).number_format = '0'
    ws4.cell(row=row_idx, column=2).number_format = '$#,##0.00'
    ws4.cell(row=row_idx, column=3).number_format = '$#,##0.00'
    ws4.cell(row=row_idx, column=4).number_format = '$#,##0.00;[Red]-$#,##0.00'
    ws4.cell(row=row_idx, column=5).number_format = '0.0%;[Red]-0.0%'
    ws4.cell(row=row_idx, column=6).number_format = '0'

    c_gain = ws4.cell(row=row_idx, column=4)
    c_gainpct = ws4.cell(row=row_idx, column=5)
    if gain > 0:
        c_gain.font = Font(name=ARIAL, size=10, bold=True, color='22543D')
        c_gainpct.font = Font(name=ARIAL, size=10, bold=True, color='22543D')
    elif gain < 0:
        c_gain.font = Font(name=ARIAL, size=10, bold=True, color='742A2A')
        c_gainpct.font = Font(name=ARIAL, size=10, bold=True, color='742A2A')

    prev_end = end_bal
    row_idx += 1

# Total row
ws4.cell(row=row_idx, column=1, value='TOTAL')
ws4.cell(row=row_idx, column=2, value=STARTING_BALANCE)
ws4.cell(row=row_idx, column=3, value=v2_summary['final_balance'])
ws4.cell(row=row_idx, column=4, value=v2_summary['final_balance'] - STARTING_BALANCE)
ws4.cell(row=row_idx, column=5, value=(v2_summary['final_balance'] - STARTING_BALANCE) / STARTING_BALANCE)
ws4.cell(row=row_idx, column=6, value=v2_summary['total_trades'])
for col_idx in range(1, 7):
    c = ws4.cell(row=row_idx, column=col_idx)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['subtotal_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['subtotal_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER
ws4.cell(row=row_idx, column=2).number_format = '$#,##0.00'
ws4.cell(row=row_idx, column=3).number_format = '$#,##0.00'
ws4.cell(row=row_idx, column=4).number_format = '$#,##0.00;[Red]-$#,##0.00'
ws4.cell(row=row_idx, column=5).number_format = '0.0%;[Red]-0.0%'
ws4.cell(row=row_idx, column=6).number_format = '0'

# Compounding projections section
proj_start_row = row_idx + 3
ws4.cell(row=proj_start_row, column=1, value='Projection — using V2 CAGR ' f"{v2_summary['cagr_pct']:.1f}% as growth rate")
ws4.cell(row=proj_start_row, column=1).font = Font(name=ARIAL, size=12, bold=True, color='2D3748')
ws4.merge_cells(start_row=proj_start_row, start_column=1, end_row=proj_start_row, end_column=6)

proj_headers = ['Starting Capital', 'After 1 Year', 'After 2 Years', 'After 3 Years', 'After 5 Years']
for col_idx, label in enumerate(proj_headers, 1):
    c = ws4.cell(row=proj_start_row + 2, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER

cagr_fraction = v2_summary['cagr_pct'] / 100
for i, start_cap in enumerate([10000, 20000, 50000, 100000]):
    r = proj_start_row + 3 + i
    vals = [
        start_cap,
        start_cap * (1 + cagr_fraction),
        start_cap * (1 + cagr_fraction)**2,
        start_cap * (1 + cagr_fraction)**3,
        start_cap * (1 + cagr_fraction)**5,
    ]
    for col_idx, val in enumerate(vals, 1):
        c = ws4.cell(row=r, column=col_idx, value=val)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        c.number_format = '$#,##0'
        c.fill = PatternFill('solid', fgColor=(THEME['row_white'] if i % 2 == 0 else THEME['row_alt']))
    ws4.cell(row=r, column=1).font = Font(name=ARIAL, size=10, bold=True, color='2D3748')

# Reality-check note
note_row = proj_start_row + 8
ws4.cell(row=note_row, column=1, value='Reality check: live WR will likely be 5-10pp lower than 69.5%. Use 50-70% of these projections as realistic outcomes. Drawdowns of 10-15% are expected even on winning strategies.')
ws4.cell(row=note_row, column=1).font = Font(name=ARIAL, size=10, italic=True, color='9B2C2C')
ws4.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
ws4.row_dimensions[note_row].height = 38

# ============================================================================
# Sheet 5: Strategy Rules (V2, with sizing methodology)
# ============================================================================
ws5 = wb.create_sheet('Strategy Rules (V2)')
ws5.merge_cells('A1:C1')
ws5['A1'] = 'SID V2 Method — Rules + Position Sizing Reference'
ws5['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws5['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 32

rules = [
    ('Stage 1 — ARM (signal day)', '', ''),
    ('Daily RSI(14) oversold (long)',    '< 30',                    'Triggers long arm'),
    ('Daily RSI(14) overbought (short)', '> 70 (V2 uses 70 not 75)','Triggers short arm. V1 used 75.'),
    ('Daily RSI(3) confirmation',        'Must be in extreme zone', 'Rebound-zone confirmation, carried from V1'),
    ('Weekly trend filter (50/200 SMA)', 'Long: 50SMA>200SMA; Short: 50SMA<200SMA', 'Retained from V1 for safety'),
    ('Earnings blackout',                '14 days pre-earnings',    'Skip new arms inside this window'),

    ('Stage 2 — TRIGGER (entry day)', '', ''),
    ('Daily RSI direction',              'Pointing in trade direction', 'Long: rising. Short: falling.'),
    ('Daily MACD direction',             'Pointing in trade direction', 'Long: rising. Short: falling.'),
    ('** Weekly RSI direction **',       'Same direction as trade (v2-weekly-or: this OR weekly MACD)', 'NEW IN V2'),
    ('** Weekly MACD direction **',      'Same direction as trade (either weekly indicator passes)', 'NEW IN V2'),
    ('** No-go zone (long) **',          'Entry RSI must be < 45',   'INSTRUCTOR S2_Ep3: "too close to RSI 50"'),
    ('** No-go zone (short) **',         'Entry RSI must be > 55',   'INSTRUCTOR S2_Ep3: "too close to RSI 50"'),
    ('Pattern bonus (W/M/H&S)',          'Bonus, NOT required',      'Deferred to V2.1 — see Pine divergence indicator'),
    ('Sticky arm window',                '3 trading days',           'If trigger not met within 3 days, arm cancels'),

    ('Stage 3 — EXIT', '', ''),
    ('Take profit',                      'Daily RSI(14) reaches 50', 'Single full exit. V1-compatible.'),
    ('Stop loss (long)',                 'Lowest low signal→entry, FLOOR to whole $', 'Hard stop, never moved'),
    ('Stop loss (short)',                'Highest high signal→entry, CEIL to whole $','Hard stop, never moved'),

    ('Risk & Position Sizing (INSTRUCTOR METHODOLOGY)', '', ''),
    ('Account balance',                  '$10,000 (this backtest)',  'Scales with growth — each trade uses current balance'),
    ('Risk per trade',                   '1% (instructor default)',  'S3_Ep4: "typically between 0.5% and 2%. We will use 1%."'),
    ('Risk $ formula',                   'Account × Risk %',         '1% of $10K = $100 risk'),
    ('Risk per share',                   '|Entry − Stop|',           'Difference in dollars per share'),
    ('Shares (sized)',                   'ROUNDDOWN(Risk $ / Risk per share)', 'Whole share count, no fractions'),
    ('Position size $',                  'Shares × Entry price',     'Capital deployed for the position'),
    ('Cash trading',                     'No margin, position ≤ account', 'Instructor teaches cash only'),
    ('Maximum risk per trade',           '2% (NEVER exceed)',        'S3_P3: "never go higher than 2%"'),
    ('When to scale up',                 'After 6-12 months live track record', 'Build confidence before adding risk'),

    ('Architecture', '', ''),
    ('SID_METHOD env var',               "Default 'v1'; set 'v2' to enable", 'Bot toggles between methods'),
    ('V1 baseline tag',                  'sid-v1-method-baseline (commit bdef9be)', 'Instant revert if V2 underperforms'),
]

start_row = 3
for r_off, (param, val, note) in enumerate(rules):
    r = start_row + r_off
    is_section = val == '' and note == ''
    is_new = '**' in param
    a = ws5.cell(row=r, column=1, value=param.replace('**', '').strip())
    b = ws5.cell(row=r, column=2, value=val)
    c = ws5.cell(row=r, column=3, value=note)
    for cell in (a, b, c):
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
        if is_section:
            cell.font = Font(name=ARIAL, size=11, bold=True, color=THEME['subheader_fg'])
            cell.fill = PatternFill('solid', fgColor=THEME['subheader_bg'])
        elif is_new:
            cell.font = Font(name=ARIAL, size=10, bold=True, color='6B46C1')
            cell.fill = PatternFill('solid', fgColor=THEME['cyan_bg'])
        else:
            cell.font = Font(name=ARIAL, size=10, color='2D3748')
            cell.fill = PatternFill('solid', fgColor=(THEME['row_alt'] if r_off % 2 == 1 else THEME['row_white']))
    if is_section:
        a.alignment = Alignment(vertical='center', horizontal='left')
    ws5.row_dimensions[r].height = 22 if is_section else 28

ws5.column_dimensions['A'].width = 38
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 55

# ============================================================================
# Sheet 6: Approval Tier Breakdown
# ============================================================================
ws6 = wb.create_sheet('Approval Tiers')
ws6.merge_cells('A1:F1')
ws6['A1'] = 'V2 Trades — Auto-Approved vs Human-Approval Required'
ws6['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws6['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws6['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 32

ws6.merge_cells('A2:F2')
ws6['A2'] = ('AUTO = ticker in proven 80-ticker tier1 list (5y V1 backtest validated). '
             'HUMAN = high-vol / crypto / leveraged ETFs / re-added dropouts. '
             'Bot sends Telegram alert with proposed Entry / SL / TP for HUMAN trades.')
ws6['A2'].font = Font(name=ARIAL, size=10, italic=True, color='2D3748')
ws6['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws6['A2'].fill = PatternFill('solid', fgColor='F7FAFC')
ws6.row_dimensions[2].height = 36

# Split trades by tier
auto_trades = [t for t in v2_sorted if t['ticker'] in AUTO_APPROVED]
human_trades = [t for t in v2_sorted if t['ticker'] not in AUTO_APPROVED]

def tier_stats(trades):
    if not trades:
        return {'n': 0, 'wins': 0, 'wr': 0, 'pnl': 0, 'avg_pnl': 0,
                'tickers': 0}
    wins = sum(1 for t in trades if t['pnl_sized'] > 0)
    pnl = sum(t['pnl_sized'] for t in trades)
    return {
        'n': len(trades),
        'wins': wins,
        'losses': len(trades) - wins,
        'wr': round(wins / len(trades) * 100, 1),
        'pnl': round(pnl, 2),
        'avg_pnl': round(pnl / len(trades), 2),
        'tickers': len(set(t['ticker'] for t in trades)),
    }

auto_stats = tier_stats(auto_trades)
human_stats = tier_stats(human_trades)

# Summary table
T_HEADER = 4
headers = ['Tier', 'Trades', 'Tickers', 'WR (%)', 'Net P&L ($)', 'Avg P&L per trade ($)']
for col_idx, label in enumerate(headers, 1):
    c = ws6.cell(row=T_HEADER, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER
    ws6.column_dimensions[get_column_letter(col_idx)].width = 22

ws6.row_dimensions[T_HEADER].height = 30

rows_data = [
    ('AUTO (tier1 80)',  auto_stats,  THEME['win_bg'],  THEME['win_fg']),
    ('HUMAN (tier2 32)', human_stats, 'FED7AA',         '9C4221'),
    ('TOTAL',            tier_stats(v2_sorted), THEME['subtotal_bg'], THEME['subtotal_fg']),
]

for i, (label, st, bg, fg) in enumerate(rows_data):
    r = T_HEADER + 1 + i
    vals = [label, st['n'], st['tickers'], st['wr'], st['pnl'], st['avg_pnl']]
    for col_idx, val in enumerate(vals, 1):
        c = ws6.cell(row=r, column=col_idx, value=val)
        c.font = Font(name=ARIAL, size=11, bold=(i == 2), color=fg)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws6.cell(row=r, column=2).number_format = '0'
    ws6.cell(row=r, column=3).number_format = '0'
    ws6.cell(row=r, column=4).number_format = '0.0'
    ws6.cell(row=r, column=5).number_format = '$#,##0.00;[Red]-$#,##0.00'
    ws6.cell(row=r, column=6).number_format = '$#,##0.00;[Red]-$#,##0.00'
    ws6.row_dimensions[r].height = 28

# Per-ticker breakdown for HUMAN tier (so user can see which need approval most)
H_HEADER = T_HEADER + 6
ws6.cell(row=H_HEADER, column=1, value='HUMAN-tier tickers — trade count, WR, P&L (rank by total trades)')
ws6.cell(row=H_HEADER, column=1).font = Font(name=ARIAL, size=12, bold=True, color='2D3748')
ws6.merge_cells(start_row=H_HEADER, start_column=1, end_row=H_HEADER, end_column=6)

h_headers = ['Ticker', 'Trades', 'Wins', 'Losses', 'WR (%)', 'Net P&L ($)']
for col_idx, label in enumerate(h_headers, 1):
    c = ws6.cell(row=H_HEADER + 2, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER

# Group by ticker
by_ticker_h = {}
for t in human_trades:
    by_ticker_h.setdefault(t['ticker'], []).append(t)
rows_h = []
for tk, ts in by_ticker_h.items():
    w = sum(1 for x in ts if x['pnl_sized'] > 0)
    pnl = sum(x['pnl_sized'] for x in ts)
    rows_h.append((tk, len(ts), w, len(ts) - w, w / len(ts) * 100, pnl))
rows_h.sort(key=lambda r: -r[1])

for i, row in enumerate(rows_h):
    r = H_HEADER + 3 + i
    tk, n, w, l, wr, pnl = row
    for col_idx, val in enumerate(row, 1):
        c = ws6.cell(row=r, column=col_idx, value=val)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        c.fill = PatternFill('solid', fgColor=(THEME['row_white'] if i % 2 == 0 else THEME['row_alt']))
    ws6.cell(row=r, column=2).number_format = '0'
    ws6.cell(row=r, column=3).number_format = '0'
    ws6.cell(row=r, column=4).number_format = '0'
    ws6.cell(row=r, column=5).number_format = '0.0'
    ws6.cell(row=r, column=6).number_format = '$#,##0.00;[Red]-$#,##0.00'
    c_pnl = ws6.cell(row=r, column=6)
    if pnl > 0:
        c_pnl.font = Font(name=ARIAL, size=10, bold=True, color='22543D')
    elif pnl < 0:
        c_pnl.font = Font(name=ARIAL, size=10, bold=True, color='742A2A')

# Save
wb.save(OUT_PATH)
print(f'\nWrote {OUT_PATH.name}')
print(f'  Sheet 1: All Trades (V2 sized) — {len(v2_sorted)} rows')
print(f'  Sheet 2: Per-Ticker Summary — {len(ticker_rows)} tickers')
print(f'  Sheet 3: V1 vs V2 (sized)')
print(f'  Sheet 4: Account Growth — equity curve + year summaries + projections')
print(f'  Sheet 5: Strategy Rules (V2) — {len(rules)} rows')
