"""
build-backtest-xlsx.py — turn SID/all-trades.csv into the formatted
"SID Strategy Back Testing.xlsx" deliverable.

Features:
  - Modern dark-teal/blue header theme (professional, not cyberpunk)
  - AutoFilter on every column so user can drill into any ticker/side/etc
  - SUBTOTAL formulas (function 9 = SUM, 3 = COUNTA) that update with filters
  - Win/Loss conditional formatting on the P&L column
  - Frozen header row
  - Per-ticker summary on a second sheet
  - Strategy parameters reference sheet
"""
import csv
import sys
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

CSV_PATH = Path(__file__).parent.parent / 'all-trades.csv'
OUT_PATH = Path(__file__).parent.parent / 'SID Strategy Back Testing.xlsx'

# ─── Theme ──────────────────────────────────────────────────────────────
THEME = {
    'title_bg':       '1A365D',  # Dark navy
    'title_fg':       'FFFFFF',
    'header_bg':      '2C5282',  # Medium blue
    'header_fg':      'FFFFFF',
    'subheader_bg':   '4299E1',  # Lighter blue (band labels)
    'subheader_fg':   'FFFFFF',
    'row_alt':        'EBF8FF',  # Very light blue
    'row_white':      'FFFFFF',
    'win_bg':         'C6F6D5',  # Light green
    'win_fg':         '22543D',
    'loss_bg':        'FED7D7',  # Light red
    'loss_fg':        '742A2A',
    'subtotal_bg':    'FEF3C7',  # Soft amber
    'subtotal_fg':    '744210',
    'border':         'CBD5E0',  # Light grey border
    'long_bg':        'BEE3F8',  # For LONG cells
    'short_bg':       'FBB6CE',  # For SHORT cells
}

THIN = Side(border_style='thin', color=THEME['border'])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ARIAL = 'Arial'

# ─── Load CSV ────────────────────────────────────────────────────────────
if not CSV_PATH.exists():
    print(f'ERROR: {CSV_PATH} not found. Run export-all-trades.py first.')
    sys.exit(1)

with open(CSV_PATH, encoding='utf-8') as f:
    trades = list(csv.DictReader(f))

print(f'Loaded {len(trades)} trades from {CSV_PATH.name}')

# ─── Build workbook ──────────────────────────────────────────────────────
wb = Workbook()

# ============================================================================
# Sheet 1: All Trades
# ============================================================================
ws = wb.active
ws.title = 'All Trades'

# Title row
TITLE = 'SID Strategy Back Testing — v1.7 with VIX >= 30 gate, 5-year window'
ws.merge_cells('A1:U1')
ws['A1'] = TITLE
ws['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

# Subtitle row with strategy + parameter summary (v1.7 shows both raw + VIX-gated)
total_pnl = sum(float(t['pnl']) for t in trades)
wins = sum(1 for t in trades if float(t['pnl']) > 0)
losses = len(trades) - wins
wr = wins / len(trades) * 100 if trades else 0

# v1.7 VIX-gate filtered stats
passed = [t for t in trades if t.get('vix_gate') == 'pass']
p_wr = sum(1 for t in passed if float(t['pnl']) > 0) / len(passed) * 100 if passed else 0
p_pnl = sum(float(t['pnl']) for t in passed)

subtitle = (
    f'{len(trades)} backtested  |  {len(passed)} after VIX>=30 production gate  |  '
    f'Win rate: {wr:.1f}% raw / {p_wr:.1f}% filtered  |  '
    f'Net P&L: ${total_pnl:+,.0f} raw / ${p_pnl:+,.0f} filtered  |  '
    f'$200 risk/trade  |  {len(set(t["ticker"] for t in trades))} ticker universe  |  '
    f'Generated {datetime.now().strftime("%Y-%m-%d")}'
)
ws.merge_cells('A2:U2')
ws['A2'] = subtitle
ws['A2'].font = Font(name=ARIAL, size=11, color='2D3748', italic=True)
ws['A2'].fill = PatternFill('solid', fgColor='F7FAFC')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# Header row (row 3)
HEADER_ROW = 3
COLS = [
    ('Trade #',          'trade_no',       'int',     8),
    ('Ticker',           'ticker',         'text',    9),
    ('Side',             'side',           'side',    9),
    ('Signal Date',      'signal_date',    'date',    13),
    ('Entry Date',       'entry_date',     'date',    13),
    ('Entry Price ($)',  'entry_price',    'money',   14),
    ('Stop Loss ($)',    'stop_loss',      'money',   12),
    ('Shares',           'shares',         'int',     9),
    ('Position ($)',     'position_value', 'money',   13),
    ('Risk ($)',         'risk_usd',       'money',   10),
    ('Exit Date',        'exit_date',      'date',    13),
    ('Exit Price ($)',   'exit_price',     'money',   13),
    ('Exit Reason',      'exit_reason',    'reason',  12),
    ('P&L ($)',          'pnl',            'pnl',     12),
    ('P&L (%)',          'pnl_pct',        'pct',     10),
    ('Bars Held',        'bars_held',      'int',     10),
    ('RSI @ Signal',     'rsi_at_signal',  'rsi',     12),
    ('RSI @ Entry',      'rsi_at_entry',   'rsi',     12),
    ('VIX @ Entry',      'vix_at_entry',   'vix',     11),
    ('VIX Gate (v1.7)',  'vix_gate',       'gate',    14),
    ('Outcome',          None,             'outcome', 11),  # computed
]

for col_idx, (label, _, _, width) in enumerate(COLS, 1):
    c = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[HEADER_ROW].height = 34

# Data rows
DATA_START = HEADER_ROW + 1
for i, t in enumerate(trades):
    row = DATA_START + i
    is_win = float(t['pnl']) > 0
    side = t['side']
    for col_idx, (_, key, kind, _) in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col_idx)
        # Value + format
        if key is None and kind == 'outcome':
            c.value = 'WIN' if is_win else 'LOSS'
        elif kind == 'int':
            c.value = int(t[key])
            c.number_format = '0'
        elif kind == 'money':
            c.value = float(t[key])
            c.number_format = '$#,##0.00;[Red]-$#,##0.00'
        elif kind == 'pnl':
            c.value = float(t[key])
            c.number_format = '$#,##0.00;[Red]-$#,##0.00'
        elif kind == 'pct':
            c.value = float(t[key]) / 100
            c.number_format = '0.00%;[Red]-0.00%'
        elif kind == 'rsi':
            c.value = float(t[key])
            c.number_format = '0.0'
        elif kind == 'vix':
            try:
                c.value = float(t[key]) if t[key] else None
                c.number_format = '0.0'
            except Exception:
                c.value = t[key]
        elif kind == 'gate':
            c.value = (t[key] or '').upper()
        elif kind == 'date':
            c.value = t[key]
            c.number_format = '@'
        else:
            c.value = t[key]
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Row-alternation
        if i % 2 == 0:
            c.fill = PatternFill('solid', fgColor=THEME['row_white'])
        else:
            c.fill = PatternFill('solid', fgColor=THEME['row_alt'])

        # Side cell colour
        if kind == 'side':
            if side == 'LONG':
                c.fill = PatternFill('solid', fgColor=THEME['long_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color='2C5282')
            else:
                c.fill = PatternFill('solid', fgColor=THEME['short_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color='9B2C2C')

        # Outcome cell colour
        if kind == 'outcome':
            if is_win:
                c.fill = PatternFill('solid', fgColor=THEME['win_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
            else:
                c.fill = PatternFill('solid', fgColor=THEME['loss_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['loss_fg'])

        # VIX Gate cell colour (PASS = green, BLOCK = red)
        if kind == 'gate':
            gate_val = (t[key] or '').lower()
            if gate_val == 'pass':
                c.fill = PatternFill('solid', fgColor=THEME['win_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
            elif gate_val == 'block':
                c.fill = PatternFill('solid', fgColor=THEME['loss_bg'])
                c.font = Font(name=ARIAL, size=10, bold=True, color=THEME['loss_fg'])

        # P&L cell colour
        if kind == 'pnl':
            if is_win:
                c.font = Font(name=ARIAL, size=10, bold=True, color='22543D')
            else:
                c.font = Font(name=ARIAL, size=10, bold=True, color='742A2A')

# Subtotal row with SUBTOTAL formulas (filter-aware)
DATA_END = DATA_START + len(trades) - 1
SUBTOTAL_ROW = DATA_END + 1

# Subtotal formula function 9 = SUM (visible cells after filter)
# Function 2 = COUNT, function 3 = COUNTA
def col_letter(idx):
    return get_column_letter(idx)

def add_subtotal(col_idx, fn, fmt=None, label=None, bold=True):
    c = ws.cell(row=SUBTOTAL_ROW, column=col_idx)
    if label:
        c.value = label
    else:
        rng = f"{col_letter(col_idx)}{DATA_START}:{col_letter(col_idx)}{DATA_END}"
        c.value = f"=SUBTOTAL({fn},{rng})"
    if fmt:
        c.number_format = fmt
    c.font = Font(name=ARIAL, size=11, bold=bold, color=THEME['subtotal_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['subtotal_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER

add_subtotal(1, 3, label='SUBTOTAL ▼')  # Trade # column shows "SUBTOTAL ▼" label
add_subtotal(2, 3, '0')                  # Ticker: count visible
ws.cell(row=SUBTOTAL_ROW, column=2).value = f'=SUBTOTAL(3,B{DATA_START}:B{DATA_END})&" trades"'
# Side: leave blank with subtotal fill
add_subtotal(3, 3, label='')
# Signal date: blank
add_subtotal(4, 3, label='')
add_subtotal(5, 3, label='')
add_subtotal(6, 1, '$#,##0.00')          # Entry price: average
add_subtotal(7, 1, '$#,##0.00')          # Stop loss: average
add_subtotal(8, 9, '#,##0')              # Shares: sum
add_subtotal(9, 9, '$#,##0.00')          # Position value: sum
add_subtotal(10, 9, '$#,##0.00')         # Risk: sum
add_subtotal(11, 3, label='')
add_subtotal(12, 1, '$#,##0.00')         # Exit price: average
add_subtotal(13, 3, label='')
add_subtotal(14, 9, '$#,##0.00;[Red]-$#,##0.00')  # P&L: sum
add_subtotal(15, 1, '0.00%;[Red]-0.00%')          # P&L %: average
add_subtotal(16, 1, '0.0')               # Bars held: average
add_subtotal(17, 1, '0.0')               # RSI at signal: average
add_subtotal(18, 1, '0.0')               # RSI at entry: average
add_subtotal(19, 1, '0.0')               # VIX at entry: average
add_subtotal(20, 3, label='')            # VIX gate: blank (text)
# Outcome: blank
add_subtotal(21, 3, label='')

ws.row_dimensions[SUBTOTAL_ROW].height = 26

# Freeze top headers (row 3 stays visible while scrolling)
ws.freeze_panes = 'C4'

# Add AutoFilter on the header row
ws.auto_filter.ref = f"A{HEADER_ROW}:U{DATA_END}"

# ============================================================================
# Sheet 2: Per-Ticker Summary
# ============================================================================
ws2 = wb.create_sheet('Per-Ticker Summary')
ws2.merge_cells('A1:G1')
ws2['A1'] = 'Per-Ticker Performance Summary'
ws2['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws2['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 32

# Aggregate per ticker
by_ticker = {}
for t in trades:
    tk = t['ticker']
    if tk not in by_ticker:
        by_ticker[tk] = []
    by_ticker[tk].append(t)

ticker_rows = []
for tk, ts in by_ticker.items():
    n = len(ts)
    w = sum(1 for t in ts if float(t['pnl']) > 0)
    pnl = sum(float(t['pnl']) for t in ts)
    avg_bars = sum(int(t['bars_held']) for t in ts) / n
    ticker_rows.append((tk, n, w, n - w, w / n * 100 if n else 0, pnl, avg_bars))

ticker_rows.sort(key=lambda r: -r[5])  # sort by P&L desc

T_HEADER_ROW = 3
T_HEADERS = [
    ('Ticker', 12),
    ('Trades', 10),
    ('Wins', 9),
    ('Losses', 10),
    ('Win Rate (%)', 14),
    ('Net P&L ($)', 14),
    ('Avg Bars Held', 14),
]
for col_idx, (label, width) in enumerate(T_HEADERS, 1):
    c = ws2.cell(row=T_HEADER_ROW, column=col_idx, value=label)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['header_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['header_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[T_HEADER_ROW].height = 32

T_DATA_START = T_HEADER_ROW + 1
for i, row in enumerate(ticker_rows):
    r = T_DATA_START + i
    tk, n, w, l, wr, pnl, avg_bars = row
    for col_idx, val in enumerate(row, 1):
        c = ws2.cell(row=r, column=col_idx, value=val)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        if i % 2 == 0:
            c.fill = PatternFill('solid', fgColor=THEME['row_white'])
        else:
            c.fill = PatternFill('solid', fgColor=THEME['row_alt'])

    # Number formats
    ws2.cell(row=r, column=2).number_format = '0'
    ws2.cell(row=r, column=3).number_format = '0'
    ws2.cell(row=r, column=4).number_format = '0'
    ws2.cell(row=r, column=5).number_format = '0.0'
    ws2.cell(row=r, column=6).number_format = '$#,##0.00;[Red]-$#,##0.00'
    ws2.cell(row=r, column=7).number_format = '0.0'

    # P&L colour
    c_pnl = ws2.cell(row=r, column=6)
    if pnl > 0:
        c_pnl.font = Font(name=ARIAL, size=10, bold=True, color='22543D')
    elif pnl < 0:
        c_pnl.font = Font(name=ARIAL, size=10, bold=True, color='742A2A')

    # WR colour
    c_wr = ws2.cell(row=r, column=5)
    if wr >= 60:
        c_wr.fill = PatternFill('solid', fgColor=THEME['win_bg'])
        c_wr.font = Font(name=ARIAL, size=10, bold=True, color=THEME['win_fg'])
    elif wr < 40:
        c_wr.fill = PatternFill('solid', fgColor=THEME['loss_bg'])
        c_wr.font = Font(name=ARIAL, size=10, bold=True, color=THEME['loss_fg'])

# Totals row
T_TOTAL_ROW = T_DATA_START + len(ticker_rows)
ws2.cell(row=T_TOTAL_ROW, column=1).value = 'TOTAL'
ws2.cell(row=T_TOTAL_ROW, column=2).value = f'=SUM(B{T_DATA_START}:B{T_TOTAL_ROW-1})'
ws2.cell(row=T_TOTAL_ROW, column=3).value = f'=SUM(C{T_DATA_START}:C{T_TOTAL_ROW-1})'
ws2.cell(row=T_TOTAL_ROW, column=4).value = f'=SUM(D{T_DATA_START}:D{T_TOTAL_ROW-1})'
ws2.cell(row=T_TOTAL_ROW, column=5).value = f'=C{T_TOTAL_ROW}/B{T_TOTAL_ROW}*100'
ws2.cell(row=T_TOTAL_ROW, column=6).value = f'=SUM(F{T_DATA_START}:F{T_TOTAL_ROW-1})'
ws2.cell(row=T_TOTAL_ROW, column=7).value = f'=AVERAGE(G{T_DATA_START}:G{T_TOTAL_ROW-1})'
for col_idx in range(1, 8):
    c = ws2.cell(row=T_TOTAL_ROW, column=col_idx)
    c.font = Font(name=ARIAL, size=11, bold=True, color=THEME['subtotal_fg'])
    c.fill = PatternFill('solid', fgColor=THEME['subtotal_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER
ws2.cell(row=T_TOTAL_ROW, column=2).number_format = '0'
ws2.cell(row=T_TOTAL_ROW, column=3).number_format = '0'
ws2.cell(row=T_TOTAL_ROW, column=4).number_format = '0'
ws2.cell(row=T_TOTAL_ROW, column=5).number_format = '0.0'
ws2.cell(row=T_TOTAL_ROW, column=6).number_format = '$#,##0.00;[Red]-$#,##0.00'
ws2.cell(row=T_TOTAL_ROW, column=7).number_format = '0.0'
ws2.row_dimensions[T_TOTAL_ROW].height = 24

ws2.freeze_panes = 'A4'
ws2.auto_filter.ref = f"A{T_HEADER_ROW}:G{T_TOTAL_ROW-1}"

# ============================================================================
# Sheet 3: Strategy Parameters
# ============================================================================
ws3 = wb.create_sheet('Strategy Rules')
ws3.merge_cells('A1:C1')
ws3['A1'] = 'SID Strategy v1.5 / v1.6 Rules Used'
ws3['A1'].font = Font(name=ARIAL, size=16, bold=True, color=THEME['title_fg'])
ws3['A1'].fill = PatternFill('solid', fgColor=THEME['title_bg'])
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 32

rules = [
    ('Stage 1 — Signal (arm)', '', ''),
    ('Daily RSI(14) oversold threshold (long)', '< 30', 'RSI must drop below 30 to arm a long signal'),
    ('Daily RSI(14) overbought threshold (short)', '> 75', 'RSI must rise above 75 to arm a short signal'),
    ('Daily RSI(3) confirmation', 'Must also be in extreme zone', 'Filters out stale RSI(14) lag'),
    ('Weekly trend filter', '50-SMA > 200-SMA for long, < for short', 'Strong filter — only trade with weekly trend'),
    ('Earnings blackout', '14 days BEFORE earnings (pre-only)', 'Trading day after earnings is allowed'),
    ('PPI macro blackout (v1.6)', '14 days before US PPI release', 'Pre-only — trading day after PPI is allowed'),
    ('VIX regime gate (v1.7)', 'Skip new entries if VIX close >= 30', 'Open positions continue — gate only blocks new arms'),
    ('Stage 2 — Trigger (entry)', '', ''),
    ('Daily RSI direction', 'Pointing same way as trade', 'Long = RSI rising; Short = RSI falling'),
    ('Daily MACD direction', 'Pointing same way as trade', 'Long = MACD rising; Short = MACD falling'),
    ('Sticky arm window', '3 trading days', 'If trigger not met within 3 days, arm cancels'),
    ('Stage 3 — Exit', '', ''),
    ('Take profit', 'Daily RSI(14) = 50', 'Single full exit, no partials, no trailing'),
    ('Stop loss (long)', 'Lowest low signal->entry, FLOOR to whole $', 'Hard stop, never moved'),
    ('Stop loss (short)', 'Highest high signal->entry, CEIL to whole $', 'Hard stop, never moved'),
    ('Sizing', '', ''),
    ('Account size assumption', '$10,000', 'Position scales linearly with this'),
    ('Risk per trade', '2% ($200 fixed)', 'Each trade risks exactly this to the stop'),
    ('Position size cap', '10% of account (live bot only)', 'Backtest does NOT cap — see CSV for raw values'),
    ('Trading universe (this dataset)', '', ''),
    ('Tickers tested', '71', 'Favourites 61 + All 10 (community-sourced)'),
    ('Window', '5 years (~2021–2026)', 'Statistically meaningful sample'),
]

start_row = 3
for r_off, (param, val, note) in enumerate(rules):
    r = start_row + r_off
    is_section = val == '' and note == ''
    a = ws3.cell(row=r, column=1, value=param)
    b = ws3.cell(row=r, column=2, value=val)
    c = ws3.cell(row=r, column=3, value=note)
    for cell in (a, b, c):
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
        if is_section:
            cell.font = Font(name=ARIAL, size=11, bold=True, color=THEME['subheader_fg'])
            cell.fill = PatternFill('solid', fgColor=THEME['subheader_bg'])
        else:
            cell.font = Font(name=ARIAL, size=10, color='2D3748')
            cell.fill = PatternFill('solid', fgColor=(THEME['row_alt'] if r_off % 2 == 1 else THEME['row_white']))
    if is_section:
        a.alignment = Alignment(vertical='center', horizontal='left')
    ws3.row_dimensions[r].height = 22 if is_section else 28

ws3.column_dimensions['A'].width = 38
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 55

# Save
wb.save(OUT_PATH)
print(f'\nWrote {OUT_PATH.name}')
print(f'  Sheet 1: All Trades ({len(trades)} rows)')
print(f'  Sheet 2: Per-Ticker Summary ({len(ticker_rows)} tickers)')
print(f'  Sheet 3: Strategy Rules ({len(rules)} parameter rows)')
